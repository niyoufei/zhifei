#!/usr/bin/env python3
from __future__ import annotations

import argparse
import asyncio
import csv
import json
import math
import re
import time
from statistics import median
from pathlib import Path
from typing import Any, Dict, List, Tuple

from backend.zhifei_autoplan.parsers.boq_parser import BoQParser
from backend.zhifei_autoplan.v2.kg_retrieval_benchmark import (
    DEFAULT_DATASET_PATH as DEFAULT_BENCHMARK_DATASET_PATH,
    ensure_benchmark_dataset,
)
from backend.zhifei_autoplan.v2.kg_paths import resolve_default_kg_root
from backend.zhifei_autoplan.v2.multi_agent_pipeline import MultiAgentDocPipeline

BOQ_FILE_EXTENSIONS = {".csv", ".xlsx", ".xls", ".pdf"}
NUMERIC_LIMITS = {
    "quantity": 1.0e8,
    "unit_price": 1.0e7,
    "total_price": 1.0e12,
}
BOQ_REVIEW_QUEUE_MAX = 400
CONFIDENCE_WEIGHTS = {
    "quantity_scientific_explosion": 0.24,
    "unit_price_scientific_explosion": 0.16,
    "total_price_scientific_explosion": 0.20,
    "quantity_unit_outlier": 0.14,
    "quantity_dynamic_outlier": 0.12,
    "price_consistency_outlier": 0.12,
    "missing_unit_for_quantity": 0.08,
    "missing_price_pair": 0.08,
    "missing_price_pair_soft": 0.03,
    "missing_quantity": 0.10,
    "unrecognized_name_noise": 0.08,
    "missing_code_soft": 0.02,
}
UNIT_QUANTITY_LIMITS = {
    "m3": 1.0e7,
    "m2": 1.0e8,
    "m": 1.0e8,
    "t": 1.0e7,
    "kg": 1.0e9,
    "台": 1.0e6,
    "套": 1.0e7,
    "个": 1.0e8,
    "项": 1.0e6,
    "处": 1.0e6,
    "座": 1.0e6,
}
BOQ_HEADER_NOISE_TERMS = {
    "序号",
    "清单编码",
    "清单项目编码",
    "编码",
    "项目名称",
    "名称",
    "工程量",
    "数量",
    "单位",
    "综合单价",
    "单价",
    "合价",
    "总价",
    "金额",
    "备注",
    "页码",
    "小计",
    "合计",
}
BOQ_ROUTE_SCORE_ADJUST = {
    "boq_parser": 0.03,
    "csv": 0.02,
    "excel": 0.02,
    "pdf_table": -0.02,
    "pdf_text": -0.04,
    "pdf_ocr": -0.06,
    "unknown": -0.07,
}
BOQ_UNIT_PATTERN = r"(?:m3|m2|m|t|kg|台|套|个|项|处|座|樘|米|吨|㎡|㎥)"
RETRIEVAL_DOMAIN_CORE = {
    "bridge",
    "tunnel",
    "railway",
    "hydraulic",
    "mep",
    "earthwork",
    "road",
    "building",
    "management",
    "digital",
}
RETRIEVAL_DOMAIN_TOKEN_MAP = (
    ("bridge", "bridge"),
    ("tunnel", "tunnel"),
    ("railway", "railway"),
    ("rail", "railway"),
    ("offshorewind", "hydraulic"),
    ("marine", "hydraulic"),
    ("harbor", "hydraulic"),
    ("port", "hydraulic"),
    ("hydraulic", "hydraulic"),
    ("water", "hydraulic"),
    ("river", "hydraulic"),
    ("sponge", "hydraulic"),
    ("drainage", "hydraulic"),
    ("wtp", "hydraulic"),
    ("water-treatment", "hydraulic"),
    ("mep", "mep"),
    ("electrical", "mep"),
    ("hvac", "mep"),
    ("fire", "mep"),
    ("gas", "mep"),
    ("pipeline", "mep"),
    ("petrochemical", "mep"),
    ("power-energy", "mep"),
    ("power", "mep"),
    ("energy", "mep"),
    ("weak-current", "mep"),
    ("district-heating", "mep"),
    ("heating", "mep"),
    ("waste-to-energy", "mep"),
    ("communication", "mep"),
    ("smartsite", "digital"),
    ("smartom", "digital"),
    ("digital", "digital"),
    ("bim", "digital"),
    ("data-center", "digital"),
    ("network", "digital"),
    ("调度", "digital"),
    ("碳", "digital"),
    ("networkgraph", "digital"),
    ("quantum", "digital"),
    ("carbon", "digital"),
    ("fm", "digital"),
    ("earthwork", "earthwork"),
    ("foundation", "earthwork"),
    ("deep-excavation", "earthwork"),
    ("road", "road"),
    ("municipal-road", "road"),
    ("landscape", "road"),
    ("airport", "road"),
    ("highway", "road"),
    ("building", "building"),
    ("housing", "building"),
    ("hospital", "building"),
    ("deco", "building"),
    ("decoration", "building"),
    ("curtain", "building"),
    ("steel-structure", "building"),
    ("prefabricated", "building"),
    ("demolition", "building"),
    ("exterior-ancillary", "building"),
    ("ancillary", "building"),
    ("urban-renewal", "building"),
    ("crane", "building"),
    ("lifting", "building"),
    ("scaffolding", "building"),
    ("formwork", "building"),
    ("management", "management"),
    ("safetycivilization", "management"),
    ("greenconstruction", "management"),
    ("temporaryworks", "management"),
    ("fournew", "management"),
)


def _confidence_level(score: float) -> str:
    val = max(0.0, min(1.0, float(score)))
    if val >= 0.85:
        return "high"
    if val >= 0.65:
        return "medium"
    return "low"


def _normalize_retrieval_domain_label(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return "unknown"
    alias_map = {
        "general": "management",
        "quality": "management",
        "safety": "management",
        "environment": "management",
        "municipal": "road",
        "traffic": "road",
        "hospital": "building",
        "housing": "building",
        "decoration": "building",
        "fire": "mep",
        "electrical": "mep",
        "automation": "digital",
    }
    if text in RETRIEVAL_DOMAIN_CORE:
        return text
    if text in alias_map:
        return alias_map[text]
    if text.startswith("zf-kg-"):
        text = re.sub(r"^zf-kg-\d+-", "", text)
        text = re.sub(r"\.json$", "", text)
    compact = text.replace("_", "-").replace(" ", "-")
    for token, domain in RETRIEVAL_DOMAIN_TOKEN_MAP:
        if token in compact:
            return domain
    return "unknown"


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    text = str(value).strip().replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", text)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _sanitize_numeric(value: float | None, field: str) -> float | None:
    if value is None:
        return None
    limit = float(NUMERIC_LIMITS.get(field) or 0.0)
    if limit > 0 and abs(value) > limit:
        return None
    return value


def _normalize_unit(unit: Any) -> str:
    raw = str(unit or "").strip().lower()
    mapping = {
        "㎡": "m2",
        "m²": "m2",
        "米2": "m2",
        "米²": "m2",
        "㎡/": "m2",
        "㎥": "m3",
        "m³": "m3",
        "米3": "m3",
        "米³": "m3",
        "吨": "t",
    }
    if raw in mapping:
        return mapping[raw]
    return raw


def _quantity_limit_by_unit(unit: str) -> float:
    key = _normalize_unit(unit)
    if key in UNIT_QUANTITY_LIMITS:
        return float(UNIT_QUANTITY_LIMITS[key])
    return float(NUMERIC_LIMITS["quantity"])


def _looks_like_scientific_explosion(raw: Any) -> bool:
    text = str(raw or "").strip().lower()
    if not text:
        return False
    if "e+" in text:
        m = re.search(r"e\+(\d+)", text)
        if m and int(m.group(1)) >= 9:
            return True
    digits = re.sub(r"[^0-9]", "", text)
    return len(digits) >= 16


def _is_noise_boq_name(name: str) -> bool:
    text = str(name or "").strip()
    if not text:
        return True
    normalized = re.sub(r"\s+", "", text).lower()
    if normalized in {x.lower() for x in BOQ_HEADER_NOISE_TERMS}:
        return True
    if all(token in normalized for token in ("分部分项", "工程量清单")):
        return True
    if re.fullmatch(r"[A-Za-z]{1,5}\d{0,3}", normalized):
        return True
    if len(re.sub(r"[^A-Za-z0-9\u4e00-\u9fa5]", "", text)) < 2:
        return True
    return False


def _extract_pdf_row_from_line(clean: str, *, parse_route: str) -> Dict[str, Any] | None:
    text = re.sub(r"\s+", " ", str(clean or "").strip())
    if not text:
        return None
    qty_unit_re = re.compile(rf"(?P<qty>-?\d+(?:\.\d+)?)\s*(?P<unit>{BOQ_UNIT_PATTERN})\b", flags=re.IGNORECASE)
    m = qty_unit_re.search(text)
    if not m:
        return None
    prefix = text[: m.start()].strip(" -:：|")
    suffix = text[m.end() :].strip()
    if not prefix:
        return None

    prefix_tokens = prefix.split()
    if len(prefix_tokens) >= 2 and re.fullmatch(r"\d{1,3}", prefix_tokens[0]):
        prefix_tokens = prefix_tokens[1:]
    if not prefix_tokens:
        return None

    code = ""
    first = prefix_tokens[0]
    if re.fullmatch(r"[A-Za-z]?[A-Za-z0-9./_-]{2,24}", first):
        if re.search(r"\d", first) and len(prefix_tokens) >= 2:
            code = first
            prefix_tokens = prefix_tokens[1:]

    name = " ".join(prefix_tokens).strip()
    if _is_noise_boq_name(name):
        return None

    numeric_tail = re.findall(r"-?\d+(?:\.\d+)?(?:[eE][+-]?\d+)?", suffix or "")
    unit_price = numeric_tail[0] if len(numeric_tail) >= 1 else None
    total_price = numeric_tail[1] if len(numeric_tail) >= 2 else None
    return {
        "boq_code": code,
        "name": name,
        "quantity": m.group("qty"),
        "unit": m.group("unit"),
        "unit_price": unit_price,
        "total_price": total_price,
        "_parse_route": parse_route,
    }


def _fallback_parse_pdf_text_rows(path: Path) -> List[Dict[str, Any]]:
    try:
        import pdfplumber  # type: ignore
    except Exception:
        return []

    rows: List[Dict[str, Any]] = []
    line_re = re.compile(
        r"(?P<code>[A-Za-z0-9][A-Za-z0-9./_-]{2,})\s+"
        r"(?P<name>[\u4e00-\u9fa5A-Za-z0-9（）()\-_/·、]{2,50})\s+"
        r"(?P<qty>\d+(?:\.\d+)?)\s*"
        rf"(?P<unit>{BOQ_UNIT_PATTERN})\b",
        flags=re.IGNORECASE,
    )
    try:
        with pdfplumber.open(str(path)) as pdf:
            pages = pdf.pages[: min(20, len(pdf.pages))]
            for page in pages:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    clean = re.sub(r"\s+", " ", str(line or "").strip())
                    if not clean:
                        continue
                    parsed = _extract_pdf_row_from_line(clean, parse_route="pdf_text")
                    if parsed:
                        rows.append(parsed)
                        continue
                    m = line_re.search(clean)
                    if not m:
                        continue
                    if _is_noise_boq_name(str(m.group("name") or "")):
                        continue
                    rows.append(
                        {
                            "boq_code": m.group("code"),
                            "name": m.group("name"),
                            "quantity": m.group("qty"),
                            "unit": m.group("unit"),
                            "unit_price": None,
                            "total_price": None,
                            "_parse_route": "pdf_text",
                        }
                    )
    except Exception:
        return []
    return rows[:2000]


def _fallback_parse_pdf_ocr_rows(path: Path) -> List[Dict[str, Any]]:
    try:
        from pdf2image import convert_from_path  # type: ignore
        import pytesseract  # type: ignore
    except Exception:
        return []

    rows: List[Dict[str, Any]] = []
    line_re = re.compile(
        r"(?P<code>[A-Za-z0-9][A-Za-z0-9./_-]{2,})\s+"
        r"(?P<name>[\u4e00-\u9fa5A-Za-z0-9（）()\-_/·、]{2,50})\s+"
        r"(?P<qty>\d+(?:\.\d+)?)\s*"
        rf"(?P<unit>{BOQ_UNIT_PATTERN})\b",
        flags=re.IGNORECASE,
    )
    try:
        images = convert_from_path(str(path), dpi=170, first_page=1, last_page=8)
    except Exception:
        return []
    for image in images[:8]:
        try:
            text = pytesseract.image_to_string(image, lang="chi_sim+eng")
        except Exception:
            text = ""
        if not text:
            continue
        for line in text.splitlines():
            clean = re.sub(r"\s+", " ", str(line or "").strip())
            if not clean:
                continue
            parsed = _extract_pdf_row_from_line(clean, parse_route="pdf_ocr")
            if parsed:
                rows.append(parsed)
                if len(rows) >= 1800:
                    return rows
                continue
            m = line_re.search(clean)
            if not m:
                continue
            if _is_noise_boq_name(str(m.group("name") or "")):
                continue
            rows.append(
                {
                    "boq_code": m.group("code"),
                    "name": m.group("name"),
                    "quantity": m.group("qty"),
                    "unit": m.group("unit"),
                    "unit_price": None,
                    "total_price": None,
                    "_parse_route": "pdf_ocr",
                }
            )
            if len(rows) >= 1800:
                return rows
    return rows


def _pick_value(row: Dict[str, Any], aliases: List[str]) -> Any:
    for key in aliases:
        for row_key, row_val in row.items():
            if str(row_key).strip().lower() == key.lower():
                return row_val
    return None


def _load_boq_csv(path: Path) -> Dict[str, Any]:
    items: List[Dict[str, Any]] = []
    with path.open("r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.DictReader(f)
        for idx, row in enumerate(reader, start=1):
            name = _pick_value(row, ["name", "项目名称", "清单项目名称", "名称"])
            if not str(name or "").strip():
                continue
            item = {
                "boq_code": str(_pick_value(row, ["boq_code", "code", "清单编码", "编码", "项目编码"]) or f"CSV-{idx}"),
                "name": str(name).strip(),
                "quantity": _to_float(_pick_value(row, ["quantity", "qty", "工程量", "数量"])),
                "unit": str(_pick_value(row, ["unit", "单位", "计量单位"]) or "").strip() or None,
                "unit_price": _to_float(_pick_value(row, ["unit_price", "综合单价", "单价"])),
                "total_price": _to_float(_pick_value(row, ["total_price", "合价", "总价", "金额"])),
            }
            items.append(item)

    stats = {
        "item_count": len(items),
        "total_quantity": sum([float(it.get("quantity") or 0.0) for it in items]),
    }
    return {"items": items, "stats": stats}


def _load_boq_excel(path: Path) -> Dict[str, Any]:
    aliases = {
        "boq_code": ["boq_code", "code", "清单编码", "编码", "项目编码"],
        "name": ["name", "项目名称", "清单项目名称", "名称"],
        "quantity": ["quantity", "qty", "工程量", "数量"],
        "unit": ["unit", "单位", "计量单位"],
        "unit_price": ["unit_price", "综合单价", "单价"],
        "total_price": ["total_price", "合价", "总价", "金额"],
    }

    def _header_index(header: List[str]) -> Dict[str, int]:
        idx_map: Dict[str, int] = {}
        lower = [str(x or "").strip().lower() for x in header]
        for key, keys in aliases.items():
            for i, col in enumerate(lower):
                if any(col == k.lower() for k in keys):
                    idx_map[key] = i
                    break
        return idx_map

    rows: List[List[Any]] = []
    sheet_count = 0
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
        for ws in wb.worksheets:
            sheet_count += 1
            for r in ws.iter_rows(values_only=True):
                vals = list(r)
                if not any(str(x or "").strip() for x in vals):
                    continue
                rows.append(vals)
    except Exception:
        try:
            import pandas as pd  # type: ignore

            xls = pd.ExcelFile(str(path))
            for sheet_name in xls.sheet_names:
                sheet_count += 1
                df = pd.read_excel(xls, sheet_name=sheet_name)
                rows.append(list(df.columns))
                for _, row in df.iterrows():
                    rows.append(list(row.values))
        except Exception as exc:
            raise ValueError(f"excel parse failed: {path} | {exc}")

    if not rows:
        return {"items": [], "stats": {"item_count": 0, "sheet_count": sheet_count}}

    header_row = None
    header_map: Dict[str, int] = {}
    for i in range(min(12, len(rows))):
        test_header = [str(x or "").strip() for x in rows[i]]
        test_map = _header_index(test_header)
        if len(test_map) >= 2 and "name" in test_map:
            header_row = i
            header_map = test_map
            break
    if header_row is None:
        header_row = 0
        header_map = _header_index([str(x or "").strip() for x in rows[0]])

    items: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows[header_row + 1 :], start=1):
        get = lambda key: row[header_map[key]] if key in header_map and header_map[key] < len(row) else None
        name = get("name")
        if not str(name or "").strip():
            continue
        item = {
            "boq_code": str(get("boq_code") or f"XLS-{idx}").strip(),
            "name": str(name).strip(),
            "quantity": _to_float(get("quantity")),
            "unit": str(get("unit") or "").strip() or None,
            "unit_price": _to_float(get("unit_price")),
            "total_price": _to_float(get("total_price")),
        }
        items.append(item)

    stats = {
        "item_count": len(items),
        "total_quantity": sum([float(it.get("quantity") or 0.0) for it in items]),
        "sheet_count": sheet_count,
    }
    return {"items": items, "stats": stats}


def _fallback_parse_pdf_table_rows(path: Path) -> List[Dict[str, Any]]:
    try:
        import pdfplumber  # type: ignore
    except Exception:
        return []

    rows: List[Dict[str, Any]] = []
    try:
        with pdfplumber.open(str(path)) as pdf:
            pages = pdf.pages[: min(25, len(pdf.pages))]
            for page in pages:
                for table in page.extract_tables() or []:
                    if not isinstance(table, list) or len(table) < 2:
                        continue
                    header = [str(x or "").strip() for x in (table[0] or [])]
                    header_l = [h.lower() for h in header]
                    has_name = any(h in {"项目名称", "清单项目名称", "名称", "name"} for h in header)
                    if not has_name and not any("名称" in h for h in header):
                        continue

                    def _pick_idx(candidates: List[str], default: int | None = None) -> int | None:
                        for i, h in enumerate(header_l):
                            if any(h == c.lower() for c in candidates):
                                return i
                        for i, h in enumerate(header):
                            if any(c in str(h) for c in candidates):
                                return i
                        return default

                    idx_code = _pick_idx(["boq_code", "code", "清单编码", "编码", "项目编码"], 0)
                    idx_name = _pick_idx(["name", "项目名称", "清单项目名称", "名称"], 1)
                    idx_qty = _pick_idx(["quantity", "qty", "工程量", "数量"], 2)
                    idx_unit = _pick_idx(["unit", "单位", "计量单位"], 3)
                    idx_up = _pick_idx(["unit_price", "综合单价", "单价"])
                    idx_tp = _pick_idx(["total_price", "合价", "总价", "金额"])

                    for r in table[1:]:
                        if not isinstance(r, list):
                            continue
                        name = r[idx_name] if idx_name is not None and idx_name < len(r) else None
                        if not str(name or "").strip():
                            continue
                        rows.append(
                            {
                                "boq_code": str(r[idx_code]).strip()
                                if idx_code is not None and idx_code < len(r) and str(r[idx_code] or "").strip()
                                else "",
                                "name": str(name).strip(),
                                "quantity": r[idx_qty] if idx_qty is not None and idx_qty < len(r) else None,
                                "unit": r[idx_unit] if idx_unit is not None and idx_unit < len(r) else None,
                                "unit_price": r[idx_up] if idx_up is not None and idx_up < len(r) else None,
                                "total_price": r[idx_tp] if idx_tp is not None and idx_tp < len(r) else None,
                                "_parse_route": "pdf_table",
                            }
                        )
    except Exception:
        return []
    return rows[:3000]


def _boq_candidates(path: Path) -> List[Path]:
    if path.is_file():
        return [path]
    if not path.is_dir():
        return []
    files = [p for p in path.rglob("*") if p.is_file() and p.suffix.lower() in BOQ_FILE_EXTENSIONS]
    return sorted(files, key=lambda p: str(p))


def _normalize_boq_item(
    raw: Dict[str, Any],
    source_file: Path,
    seq: int,
    *,
    file_price_optional: bool = False,
) -> Dict[str, Any]:
    name = str(raw.get("name") or "").strip()
    if not name or _is_noise_boq_name(name):
        return {}
    anomalies: List[str] = []
    raw_unit = str(raw.get("unit") or "").strip()
    norm_unit = _normalize_unit(raw_unit)
    unit = norm_unit or raw_unit or None
    quantity_raw = raw.get("quantity")
    quantity = _sanitize_numeric(_to_float(quantity_raw), "quantity")
    if _looks_like_scientific_explosion(quantity_raw):
        quantity = None
        anomalies.append("quantity_scientific_explosion")
    if quantity is not None:
        q_limit = _quantity_limit_by_unit(unit or "")
        if abs(float(quantity)) > q_limit:
            quantity = None
            anomalies.append("quantity_unit_outlier")
    if quantity is None:
        return {}

    unit_price_raw = raw.get("unit_price")
    total_price_raw = raw.get("total_price")
    unit_price = _sanitize_numeric(_to_float(unit_price_raw), "unit_price")
    total_price = _sanitize_numeric(_to_float(total_price_raw), "total_price")
    if _looks_like_scientific_explosion(unit_price_raw):
        unit_price = None
        anomalies.append("unit_price_scientific_explosion")
    if _looks_like_scientific_explosion(total_price_raw):
        total_price = None
        anomalies.append("total_price_scientific_explosion")
    if total_price is None and quantity is not None and unit_price is not None:
        total_price = _sanitize_numeric(quantity * unit_price, "total_price")
    if (
        total_price is not None
        and quantity is not None
        and unit_price is not None
        and abs(float(quantity) * float(unit_price)) > 0
    ):
        expect = abs(float(quantity) * float(unit_price))
        ratio = abs(float(total_price) - expect) / max(expect, 1.0)
        if ratio > 8.0:
            anomalies.append("price_consistency_outlier")
    # Rows without any numeric signal are very likely OCR/header noise and should not enter governance.
    if quantity is None and unit_price is None and total_price is None:
        return {}
    if quantity is not None and (unit is None or not str(unit).strip()):
        anomalies.append("missing_unit_for_quantity")
    if unit_price is None and total_price is None:
        parse_route_for_price = str(raw.get("_parse_route") or "unknown").strip().lower()
        if file_price_optional or parse_route_for_price in {"pdf_text", "pdf_ocr"}:
            anomalies.append("missing_price_pair_soft")
        else:
            anomalies.append("missing_price_pair")

    confidence = 1.0
    for tag in anomalies:
        confidence -= float(CONFIDENCE_WEIGHTS.get(str(tag), 0.06))
    parse_route = str(raw.get("_parse_route") or "unknown").strip().lower()
    confidence += float(BOQ_ROUTE_SCORE_ADJUST.get(parse_route, BOQ_ROUTE_SCORE_ADJUST["unknown"]))

    code_text = str(raw.get("boq_code") or raw.get("code") or "").strip()
    if code_text:
        if re.search(r"[A-Za-z]", code_text) and re.search(r"\d", code_text):
            confidence += 0.02
    else:
        anomalies.append("missing_code_soft")
        confidence -= float(CONFIDENCE_WEIGHTS.get("missing_code_soft", 0.02))
    if code_text.startswith(("AUTO-", "CSV-", "XLS-")):
        confidence -= 0.03

    if quantity is not None and unit is not None:
        confidence += 0.06
    if unit_price is not None or total_price is not None:
        confidence += 0.03
    resources = raw.get("resources")
    process_name = str(raw.get("process") or "").strip()
    if process_name:
        confidence += 0.02
    if isinstance(resources, list) and resources:
        confidence += 0.02
    if isinstance(raw.get("source_file"), str) and str(raw.get("source_file") or "").strip():
        confidence += 0.01

    confidence = max(0.0, min(1.0, confidence))

    return {
        "boq_code": str(code_text or f"AUTO-{seq}"),
        "name": name,
        "quantity": quantity,
        "unit": unit,
        "unit_price": unit_price,
        "total_price": total_price,
        "source_file": str(source_file),
        "parse_route": parse_route,
        "parsing_confidence": round(confidence, 6),
        "confidence_level": _confidence_level(confidence),
        "anomalies": anomalies,
    }


async def _load_single_boq(path: Path) -> Dict[str, Any]:
    suffix = path.suffix.lower()
    filename = str(path.name or "")
    price_optional = bool(
        suffix == ".pdf"
        and any(token in filename for token in ("汇总", "透视", "工程量清单", "清单汇总", "清单透视", "清单总表"))
    )
    fallback_used = False
    fallback_sources: List[str] = []
    if suffix == ".csv":
        payload = _load_boq_csv(path)
    elif suffix in {".xlsx", ".xls"}:
        payload = _load_boq_excel(path)
    else:
        parser = BoQParser()
        items, stats = await parser.parse(str(path))
        parser_rows: List[Dict[str, Any]] = []
        for item in items:
            row = item.model_dump() if hasattr(item, "model_dump") else dict(item)
            row["_parse_route"] = "boq_parser"
            parser_rows.append(row)
        payload = {
            "items": parser_rows,
            "stats": stats,
        }
        if suffix == ".pdf":
            parser_rows = list(payload.get("items") or [])
            table_rows = _fallback_parse_pdf_table_rows(path)
            text_rows = _fallback_parse_pdf_text_rows(path)
            ocr_rows = _fallback_parse_pdf_ocr_rows(path)
            merged_rows: List[Dict[str, Any]] = []
            if parser_rows:
                merged_rows.extend(parser_rows)
            if table_rows:
                merged_rows.extend(table_rows)
                fallback_sources.append("pdf_table")
            if text_rows:
                merged_rows.extend(text_rows)
                fallback_sources.append("pdf_text")
            if ocr_rows:
                merged_rows.extend(ocr_rows)
                fallback_sources.append("pdf_ocr")
            if merged_rows:
                payload["items"] = merged_rows
                payload["stats"] = dict(payload.get("stats") or {})
                payload["stats"]["fallback_table_rows"] = len(table_rows)
                payload["stats"]["fallback_text_rows"] = len(text_rows)
                payload["stats"]["fallback_ocr_rows"] = len(ocr_rows)
                payload["stats"]["fusion_sources"] = ["boq_parser"] + fallback_sources
                fallback_used = bool(fallback_sources)
    normalized_items: List[Dict[str, Any]] = []
    anomaly_items: List[Dict[str, Any]] = []
    for idx, item in enumerate(payload.get("items") or [], start=1):
        if not isinstance(item, dict):
            continue
        normalized = _normalize_boq_item(item, path, idx, file_price_optional=price_optional)
        if normalized:
            normalized_items.append(normalized)
            significant_anomalies = [
                str(tag) for tag in (normalized.get("anomalies") or []) if str(tag) not in {"missing_price_pair_soft", "missing_code_soft"}
            ]
            if significant_anomalies:
                anomaly_items.append(
                    {
                        "boq_code": normalized.get("boq_code"),
                        "name": normalized.get("name"),
                        "anomalies": significant_anomalies,
                    }
                )

    quantities = [float(it.get("quantity")) for it in normalized_items if isinstance(it.get("quantity"), (int, float))]
    if quantities:
        med = float(median(quantities))
        dynamic_limit = max(_quantity_limit_by_unit(""), med * 2000.0)
        for it in normalized_items:
            q = it.get("quantity")
            if not isinstance(q, (int, float)):
                continue
            if abs(float(q)) > dynamic_limit:
                it["quantity"] = None
                anomalies = list(it.get("anomalies") or [])
                if "quantity_dynamic_outlier" not in anomalies:
                    anomalies.append("quantity_dynamic_outlier")
                it["anomalies"] = anomalies
                anomaly_items.append(
                    {
                        "boq_code": it.get("boq_code"),
                        "name": it.get("name"),
                        "anomalies": ["quantity_dynamic_outlier"],
                    }
                )

    payload["items"] = normalized_items
    payload["source_file"] = str(path)
    payload["stats"] = dict(payload.get("stats") or {})
    payload["stats"]["item_count"] = len(normalized_items)
    payload["stats"]["total_quantity"] = sum(float(it.get("quantity") or 0.0) for it in normalized_items)
    payload["stats"]["anomaly_count"] = len(anomaly_items)
    payload["stats"]["anomaly_items"] = anomaly_items[:120]
    payload["stats"]["fallback_used"] = bool(fallback_used)
    payload["stats"]["fallback_sources"] = list(fallback_sources)
    payload["stats"]["valid_quantity_count"] = sum(
        1 for it in normalized_items if isinstance(it.get("quantity"), (int, float))
    )
    payload["stats"]["valid_price_count"] = sum(
        1
        for it in normalized_items
        if isinstance(it.get("unit_price"), (int, float)) or isinstance(it.get("total_price"), (int, float))
    )
    payload["stats"]["valid_unit_count"] = sum(1 for it in normalized_items if str(it.get("unit") or "").strip())
    payload["stats"]["avg_parsing_confidence"] = round(
        sum(float(it.get("parsing_confidence") or 0.0) for it in normalized_items) / max(len(normalized_items), 1),
        6,
    )
    payload["stats"]["low_confidence_count"] = sum(
        1 for it in normalized_items if float(it.get("parsing_confidence") or 0.0) < 0.55
    )
    payload["stats"]["medium_confidence_count"] = sum(
        1 for it in normalized_items if 0.55 <= float(it.get("parsing_confidence") or 0.0) < 0.85
    )
    payload["stats"]["high_confidence_count"] = sum(
        1 for it in normalized_items if float(it.get("parsing_confidence") or 0.0) >= 0.85
    )
    route_dist: Dict[str, int] = {}
    for it in normalized_items:
        route = str(it.get("parse_route") or "unknown")
        route_dist[route] = int(route_dist.get(route) or 0) + 1
    payload["stats"]["parse_route_distribution"] = route_dist
    payload["stats"]["file_price_optional"] = bool(price_optional)
    payload["stats"]["low_confidence_items"] = [
        {
            "boq_code": str(it.get("boq_code") or ""),
            "name": str(it.get("name") or ""),
            "parsing_confidence": round(float(it.get("parsing_confidence") or 0.0), 6),
            "parse_route": str(it.get("parse_route") or ""),
            "anomalies": [str(x) for x in (it.get("anomalies") or [])[:4]],
        }
        for it in normalized_items
        if float(it.get("parsing_confidence") or 0.0) < 0.55
    ][:120]
    if not payload.get("items"):
        raise ValueError(f"BOQ parsing returned empty items: {path}")
    return payload


def _dedupe_boq_items(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    deduped: List[Dict[str, Any]] = []
    seen = set()
    for item in items:
        key = (
            str(item.get("boq_code") or "").strip().lower(),
            str(item.get("name") or "").strip().lower(),
            float(item.get("quantity") or 0.0),
            str(item.get("unit") or "").strip().lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(item)
    return deduped


async def _load_boq_payload(path: Path) -> Dict[str, Any]:
    candidates = _boq_candidates(path)
    if not candidates:
        raise ValueError(f"No BOQ files found under: {path}")

    merged_items: List[Dict[str, Any]] = []
    parse_errors: List[Dict[str, str]] = []
    file_item_count: Dict[str, int] = {}
    source_stats: Dict[str, Any] = {}
    anomaly_rows: List[Dict[str, Any]] = []

    for file_path in candidates:
        try:
            payload = await _load_single_boq(file_path)
            items = payload.get("items") or []
            merged_items.extend(items)
            file_item_count[str(file_path)] = len(items)
            source_stats[str(file_path)] = payload.get("stats") or {}
            for row in (payload.get("stats") or {}).get("anomaly_items") or []:
                if isinstance(row, dict):
                    anomaly_rows.append({**row, "source_file": str(file_path)})
        except Exception as exc:
            parse_errors.append({"file": str(file_path), "error": str(exc)})

    merged_items = _dedupe_boq_items(merged_items)
    if not merged_items:
        raise ValueError(f"BOQ parsing returned empty items for all candidates: {path}; errors={parse_errors}")

    top_quantity_items = sorted(
        [it for it in merged_items if isinstance(it.get("quantity"), (int, float))],
        key=lambda it: float(it.get("quantity") or 0.0),
        reverse=True,
    )[:10]

    return {
        "items": merged_items,
        "stats": {
            "item_count": len(merged_items),
            "total_quantity": sum(float(it.get("quantity") or 0.0) for it in merged_items),
            "source_file_count": len(candidates),
            "parsed_file_count": len(file_item_count),
            "failed_file_count": len(parse_errors),
            "file_item_count": file_item_count,
            "source_stats": source_stats,
            "top_quantity_items": top_quantity_items,
            "anomaly_count": len(anomaly_rows),
            "anomaly_items": anomaly_rows[:300],
        },
        "source_files": [str(p) for p in candidates],
        "parse_errors": parse_errors,
    }


def _build_boq_governance(
    *,
    boq_payload: Dict[str, Any],
    trust_threshold: float,
) -> Dict[str, Any]:
    stats = boq_payload.get("stats") if isinstance(boq_payload.get("stats"), dict) else {}
    source_stats = stats.get("source_stats") if isinstance(stats.get("source_stats"), dict) else {}
    parse_errors = boq_payload.get("parse_errors") if isinstance(boq_payload.get("parse_errors"), list) else []
    file_item_count = stats.get("file_item_count") if isinstance(stats.get("file_item_count"), dict) else {}

    file_scores: List[Dict[str, Any]] = []
    weighted_sum = 0.0
    weighted_count = 0.0
    manual_review_queue: List[Dict[str, Any]] = []

    for file_path, row in source_stats.items():
        info = row if isinstance(row, dict) else {}
        item_count = int(info.get("item_count") or file_item_count.get(file_path) or 0)
        anomaly_count = int(info.get("anomaly_count") or 0)
        valid_qty = int(info.get("valid_quantity_count") or 0)
        valid_unit = int(info.get("valid_unit_count") or 0)
        valid_price = int(info.get("valid_price_count") or 0)
        fallback_used = bool(info.get("fallback_used"))
        avg_conf = max(0.0, min(1.0, float(info.get("avg_parsing_confidence") or 0.0)))
        anomaly_ratio = float(anomaly_count) / max(item_count, 1)
        qty_ratio = float(valid_qty) / max(item_count, 1)
        unit_ratio = float(valid_unit) / max(item_count, 1)
        price_ratio = float(valid_price) / max(item_count, 1)
        low_conf_ratio = float(info.get("low_confidence_count") or 0) / max(item_count, 1)
        route_dist = info.get("parse_route_distribution") if isinstance(info.get("parse_route_distribution"), dict) else {}
        ocr_ratio = float(route_dist.get("pdf_ocr") or 0) / max(item_count, 1)
        text_ratio = float(route_dist.get("pdf_text") or 0) / max(item_count, 1)
        fallback_penalty = 0.0
        if fallback_used:
            fallback_penalty = min(0.06, ocr_ratio * 0.04 + text_ratio * 0.03 + 0.02)
        low_conf_penalty = min(0.10, low_conf_ratio * 0.10)
        score = max(
            0.0,
            min(
                1.0,
                avg_conf * 0.60
                + max(0.0, 1.0 - anomaly_ratio) * 0.12
                + qty_ratio * 0.14
                + unit_ratio * 0.09
                + price_ratio * 0.05
                - fallback_penalty,
            ),
        )
        score = max(0.0, min(1.0, score - low_conf_penalty))
        trust_level = "high" if score >= 0.85 else "medium" if score >= trust_threshold else "low"
        file_scores.append(
            {
                "file": str(file_path),
                "item_count": item_count,
                "anomaly_count": anomaly_count,
                "anomaly_ratio": round(anomaly_ratio, 6),
                "valid_quantity_ratio": round(qty_ratio, 6),
                "valid_unit_ratio": round(unit_ratio, 6),
                "valid_price_ratio": round(price_ratio, 6),
                "avg_parsing_confidence": round(avg_conf, 6),
                "low_confidence_ratio": round(low_conf_ratio, 6),
                "fallback_used": fallback_used,
                "trust_score": round(score, 6),
                "trust_level": trust_level,
            }
        )
        weighted_sum += score * max(item_count, 1)
        weighted_count += max(item_count, 1)

        if trust_level == "low" or anomaly_count > max(12, int(math.ceil(item_count * 0.12))):
            for row_item in (info.get("anomaly_items") or [])[:80]:
                if isinstance(row_item, dict):
                    manual_review_queue.append(
                        {
                            "file": str(file_path),
                            "type": "anomaly_item",
                            "boq_code": str(row_item.get("boq_code") or ""),
                            "name": str(row_item.get("name") or ""),
                            "anomalies": [str(x) for x in (row_item.get("anomalies") or [])[:6]],
                        }
                    )
            for row_item in (info.get("low_confidence_items") or [])[:80]:
                if isinstance(row_item, dict):
                    manual_review_queue.append(
                        {
                            "file": str(file_path),
                            "type": "low_confidence_item",
                            "boq_code": str(row_item.get("boq_code") or ""),
                            "name": str(row_item.get("name") or ""),
                            "parsing_confidence": float(row_item.get("parsing_confidence") or 0.0),
                            "parse_route": str(row_item.get("parse_route") or ""),
                            "anomalies": [str(x) for x in (row_item.get("anomalies") or [])[:6]],
                        }
                    )

    for err in parse_errors:
        if not isinstance(err, dict):
            continue
        manual_review_queue.append(
            {
                "file": str(err.get("file") or ""),
                "type": "parse_error",
                "error": str(err.get("error") or "")[:400],
            }
        )

    overall = round(weighted_sum / max(weighted_count, 1.0), 6)
    low_trust_files = [x for x in file_scores if str(x.get("trust_level")) == "low"]
    source_file_total = max(int(stats.get("source_file_count") or len(file_scores) or 1), 1)
    parse_error_rate = round(len(parse_errors) / source_file_total, 6)
    low_trust_ratio = round(len(low_trust_files) / source_file_total, 6)
    trusted = bool(
        overall >= trust_threshold
        and parse_error_rate <= 0.35
        and low_trust_ratio <= 0.34
    )
    governance = {
        "enabled": True,
        "trust_threshold": round(float(trust_threshold), 6),
        "overall_trust_score": overall,
        "trusted": trusted,
        "parse_error_rate": parse_error_rate,
        "low_trust_ratio": low_trust_ratio,
        "source_file_total": source_file_total,
        "file_scores": sorted(file_scores, key=lambda x: (x.get("trust_score", 0.0), x.get("file", ""))),
        "low_trust_files": low_trust_files,
        "manual_review_queue": manual_review_queue[:BOQ_REVIEW_QUEUE_MAX],
        "manual_review_total": len(manual_review_queue),
        "trusted_with_watchlist": bool(trusted and len(low_trust_files) > 0),
        "hard_gate_recommended": bool((not trusted) or parse_error_rate > 0.35),
    }
    return governance


def _write_boq_manual_review_report(governance: Dict[str, Any], *, output_path: Path | str) -> str:
    out = Path(output_path).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    lines: List[str] = []
    lines.append("# BOQ Manual Review Queue")
    lines.append("")
    lines.append(f"- Generated At: {time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())}")
    lines.append(f"- Trusted: {bool(governance.get('trusted'))}")
    lines.append(f"- Overall Trust Score: {float(governance.get('overall_trust_score') or 0.0):.4f}")
    lines.append(f"- Parse Error Rate: {float(governance.get('parse_error_rate') or 0.0):.4f}")
    lines.append(f"- Low Trust File Ratio: {float(governance.get('low_trust_ratio') or 0.0):.4f}")
    lines.append(f"- Hard Gate Recommended: {bool(governance.get('hard_gate_recommended'))}")
    lines.append("")
    lines.append("## Low Trust Files")
    lines.append("")
    low_files = governance.get("low_trust_files") if isinstance(governance.get("low_trust_files"), list) else []
    if not low_files:
        lines.append("None")
    else:
        lines.append("| File | Trust Score | Item Count | Anomaly Count | Avg Confidence | Low Confidence Ratio |")
        lines.append("|---|---:|---:|---:|---:|---:|")
        for row in low_files:
            if not isinstance(row, dict):
                continue
            lines.append(
                f"| {row.get('file')} | {float(row.get('trust_score') or 0.0):.4f} | "
                f"{int(row.get('item_count') or 0)} | {int(row.get('anomaly_count') or 0)} | "
                f"{float(row.get('avg_parsing_confidence') or 0.0):.4f} | "
                f"{float(row.get('low_confidence_ratio') or 0.0):.4f} |"
            )
    lines.append("")
    lines.append("## Manual Review Items")
    lines.append("")
    queue = governance.get("manual_review_queue") if isinstance(governance.get("manual_review_queue"), list) else []
    if not queue:
        lines.append("None")
    else:
        lines.append("| # | File | Type | BOQ Code | Name | Details |")
        lines.append("|---|---|---|---|---|---|")
        for idx, row in enumerate(queue, start=1):
            if not isinstance(row, dict):
                continue
            details = str(row.get("error") or "、".join([str(x) for x in (row.get("anomalies") or [])]))
            conf = row.get("parsing_confidence")
            if conf is not None:
                details = f"{details}; conf={float(conf):.4f}; route={row.get('parse_route')}"
            details = details.replace("|", "\\|")
            lines.append(
                f"| {idx} | {row.get('file')} | {row.get('type')} | {row.get('boq_code') or ''} | "
                f"{row.get('name') or ''} | {details} |"
            )
    lines.append("")
    out.write_text("\n".join(lines), encoding="utf-8")
    return str(out)


def _build_retrieval_remediation_plan(
    *,
    benchmark: Dict[str, Any],
    domain_warnings: List[Dict[str, Any]],
    quality_warnings: List[Dict[str, Any]],
    min_domain_pass_rate: float,
) -> Dict[str, Any]:
    rows = benchmark.get("rows") if isinstance(benchmark.get("rows"), list) else []
    floor = max(0.0, min(float(min_domain_pass_rate), 1.0))
    weak_domains: Dict[str, Dict[str, Any]] = {}
    for row in domain_warnings:
        if not isinstance(row, dict):
            continue
        domain = str(row.get("domain") or "").strip()
        if not domain:
            continue
        weak_domains[domain] = {
            "domain": domain,
            "raw_domain": str(row.get("raw_domain") or domain),
            "total_cases": int(row.get("total_cases") or 0),
            "pass_rate": float(row.get("pass_rate") or 0.0),
            "min_pass_rate": float(row.get("min_pass_rate") or floor),
            "failed_queries": [],
            "candidate_keywords": [],
        }

    token_freq: Dict[str, Dict[str, int]] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        ok = bool(row.get("ok"))
        raw_domain = str(row.get("domain") or "").strip()
        norm_domain = _normalize_retrieval_domain_label(raw_domain)
        if norm_domain not in weak_domains:
            continue
        query = str(row.get("query") or "").strip()
        if (not ok) and query:
            weak_domains[norm_domain]["failed_queries"].append(query)
        kws = row.get("expected_keywords") if isinstance(row.get("expected_keywords"), list) else []
        bag = token_freq.setdefault(norm_domain, {})
        for kw in kws:
            term = str(kw or "").strip()
            if len(term) < 2:
                continue
            bag[term] = int(bag.get(term) or 0) + 1
        if query:
            for tok in re.split(r"[^\w\u4e00-\u9fff]+", query):
                term = str(tok or "").strip()
                if len(term) < 2:
                    continue
                bag[term] = int(bag.get(term) or 0) + 1

    actionable: List[Dict[str, Any]] = []
    for domain, row in weak_domains.items():
        bag = token_freq.get(domain) or {}
        ranked = sorted(bag.items(), key=lambda x: (-x[1], x[0]))
        candidate_keywords = [x[0] for x in ranked[:8]]
        failed_queries = [str(x) for x in row.get("failed_queries", []) if str(x).strip()][:8]
        row["candidate_keywords"] = candidate_keywords
        row["failed_queries"] = failed_queries
        row["suggested_actions"] = [
            f"补充 {domain} 领域参数节点（动作+参数+检查人）",
            "补充至少2条可计算 FormulaNode（含变量定义与单位）",
            "为新增节点补齐 reference_standard 与 evidence_sources",
            f"复跑检索评测，确保 pass_rate >= {row.get('min_pass_rate'):.2f}",
        ]
        actionable.append(row)

    quality_rows: List[Dict[str, Any]] = []
    for row in quality_warnings:
        if not isinstance(row, dict):
            continue
        quality_rows.append(
            {
                "raw_domain": str(row.get("raw_domain") or "unknown"),
                "total_cases": int(row.get("total_cases") or 0),
                "pass_rate": float(row.get("pass_rate") or 0.0),
                "suggested_actions": [
                    "清洗评测集 domain_hint，改为 canonical domain（building/road/mep/...)",
                    "避免将文件名直接作为 domain，改用工程专业域标签",
                ],
            }
        )

    return {
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
        "benchmark_ok": bool(benchmark.get("ok")),
        "total_cases": int(benchmark.get("total_cases") or 0),
        "pass_rate": float(benchmark.get("pass_rate") or 0.0),
        "avg_mrr": float(benchmark.get("avg_mrr") or 0.0),
        "domain_warnings_total": len(actionable),
        "domain_quality_warnings_total": len(quality_rows),
        "domain_actions": actionable,
        "domain_quality_actions": quality_rows,
    }


def _write_retrieval_remediation_report(plan: Dict[str, Any], *, output_path: Path | str) -> str:
    out = Path(output_path).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    lines: List[str] = []
    lines.append("# KG Retrieval Remediation Report")
    lines.append("")
    lines.append(f"- Generated At: {plan.get('generated_at')}")
    lines.append(f"- Benchmark OK: {bool(plan.get('benchmark_ok'))}")
    lines.append(f"- Total Cases: {int(plan.get('total_cases') or 0)}")
    lines.append(f"- Pass Rate: {float(plan.get('pass_rate') or 0.0):.4f}")
    lines.append(f"- Avg MRR: {float(plan.get('avg_mrr') or 0.0):.4f}")
    lines.append(f"- Domain Warnings: {int(plan.get('domain_warnings_total') or 0)}")
    lines.append(f"- Domain Label Quality Warnings: {int(plan.get('domain_quality_warnings_total') or 0)}")
    lines.append("")
    lines.append("## Domain Remediation")
    lines.append("")
    actions = plan.get("domain_actions") if isinstance(plan.get("domain_actions"), list) else []
    if not actions:
        lines.append("None")
    else:
        for idx, row in enumerate(actions, start=1):
            if not isinstance(row, dict):
                continue
            lines.append(f"### {idx}. {row.get('domain')}")
            lines.append("")
            lines.append(f"- Raw Domain: {row.get('raw_domain')}")
            lines.append(f"- Cases: {int(row.get('total_cases') or 0)}")
            lines.append(f"- Pass Rate: {float(row.get('pass_rate') or 0.0):.4f}")
            lines.append(f"- Target Pass Rate: {float(row.get('min_pass_rate') or 0.0):.4f}")
            kws = [str(x) for x in (row.get("candidate_keywords") or []) if str(x).strip()]
            lines.append(f"- Candidate Keywords: {', '.join(kws) if kws else '-'}")
            fq = [str(x) for x in (row.get("failed_queries") or []) if str(x).strip()]
            if fq:
                lines.append("- Failed Query Samples:")
                for q in fq[:6]:
                    lines.append(f"  - {q}")
            for act in (row.get("suggested_actions") or [])[:6]:
                lines.append(f"- Action: {act}")
            lines.append("")
    lines.append("## Domain Label Data Quality")
    lines.append("")
    quality_rows = plan.get("domain_quality_actions") if isinstance(plan.get("domain_quality_actions"), list) else []
    if not quality_rows:
        lines.append("None")
    else:
        lines.append("| Raw Domain | Cases | Pass Rate | Suggested Action |")
        lines.append("|---|---:|---:|---|")
        for row in quality_rows:
            if not isinstance(row, dict):
                continue
            actions_text = "；".join([str(x) for x in (row.get("suggested_actions") or [])[:2]])
            lines.append(
                f"| {row.get('raw_domain')} | {int(row.get('total_cases') or 0)} | "
                f"{float(row.get('pass_rate') or 0.0):.4f} | {actions_text} |"
            )
    lines.append("")
    out.write_text("\n".join(lines), encoding="utf-8")
    return str(out)


def _arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Run V2 engine on a real project (production mode).")
    p.add_argument("--tender", nargs="+", required=True, help="招标文件路径（PDF/Word/TXT等）。")
    p.add_argument("--boq", required=True, help="工程量清单路径（Excel/CSV/PDF 文件，或目录自动合并解析）。")
    p.add_argument(
        "--kg-root",
        default=str(resolve_default_kg_root()),
        help="知识图谱根目录。",
    )
    p.add_argument(
        "--kg-db",
        default="backend/data/autoplan/v2/knowledge_graph.sqlite3",
        help="图谱SQLite索引路径。",
    )
    p.add_argument(
        "--out",
        default="build/real_project_diagnosis.json",
        help="诊断JSON输出路径（不导出Word/PDF）。",
    )
    p.add_argument(
        "--missing-report",
        default="build/Missing_Knowledge_Report.md",
        help="知识盲区体检报告输出路径。",
    )
    p.add_argument(
        "--docx-out",
        default="/Users/youfeini/Desktop/文档生成系统/01_真实项目测试/最终施组草案_带AI审校标记.docx",
        help="最终施组草案DOCX输出路径。",
    )
    p.add_argument(
        "--docx-export",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="是否导出最终DOCX文档。",
    )
    p.add_argument(
        "--self-heal",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="启用知识图谱自愈合Agent并在缺口后自动二次重跑。",
    )
    p.add_argument(
        "--self-heal-provider",
        default=None,
        help="自愈Agent模型提供商（默认自动选择）。",
    )
    p.add_argument(
        "--self-heal-model",
        default=None,
        help="自愈Agent模型名称（默认自动选择）。",
    )
    p.add_argument(
        "--standard-auto-update",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="是否在运行前执行标准自动更新检查。",
    )
    p.add_argument(
        "--retrieval-benchmark-gate",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="是否执行检索评测集门禁。",
    )
    p.add_argument(
        "--benchmark-dataset",
        default=str(DEFAULT_BENCHMARK_DATASET_PATH),
        help="检索评测集JSON路径。",
    )
    p.add_argument(
        "--benchmark-min-pass-rate",
        type=float,
        default=0.85,
        help="检索门禁最小通过率阈值。",
    )
    p.add_argument(
        "--benchmark-min-avg-mrr",
        type=float,
        default=0.65,
        help="检索门禁最小平均MRR阈值。",
    )
    p.add_argument(
        "--benchmark-min-domain-pass-rate",
        type=float,
        default=0.70,
        help="检索门禁各专业域最小通过率阈值。",
    )
    p.add_argument(
        "--benchmark-domain-min-cases",
        type=int,
        default=3,
        help="触发域级门禁判定的最小案例数。",
    )
    p.add_argument(
        "--retrieval-remediation-report",
        default="build/KG_Retrieval_Remediation_Report.md",
        help="检索域补强计划报告输出路径。",
    )
    p.add_argument(
        "--enforce-retrieval-gate",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="检索门禁失败时是否拦截发布。",
    )
    p.add_argument(
        "--enforce-benchmark-domain-gate",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="是否启用检索评测域级门禁（按专业域逐项放行）。",
    )
    p.add_argument(
        "--feedback-learning",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="是否启用真实项目回灌学习。",
    )
    p.add_argument(
        "--feedback-writeback",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="是否将反馈学习结果回写到知识图谱节点在线学习画像。",
    )
    p.add_argument(
        "--retrieval-weight-training",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="是否启用检索权重训练闭环。",
    )
    p.add_argument(
        "--retrieval-weight-profile",
        default="build/kg_retrieval_weight_profile.json",
        help="检索权重配置文件输出路径。",
    )
    p.add_argument(
        "--region-context",
        default=None,
        help="区域上下文（如 CN/SH/BJ），用于地域法规策略筛选。",
    )
    p.add_argument(
        "--bid-date",
        default=time.strftime("%Y-%m-%d", time.localtime()),
        help="投标日期（YYYY-MM-DD），用于标准时间窗生效过滤。",
    )
    p.add_argument(
        "--allow-superseded",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="是否允许检索命中已被替代的标准节点。",
    )
    p.add_argument(
        "--regional-plugin-dir",
        default=None,
        help="地域法规插件目录（JSON）。",
    )
    p.add_argument(
        "--prefer-human-verified-hits",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="检索时优先人工/已校核节点。",
    )
    p.add_argument(
        "--min-source-weight",
        type=int,
        default=1,
        help="检索节点最小效力层级权重（答疑5>图纸4>国标3>行标2>企标1>未知0）。",
    )
    p.add_argument(
        "--release-freeze",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="通过后是否创建知识图谱冻结快照。",
    )
    p.add_argument(
        "--release-approver",
        default="system",
        help="审批与冻结签署人。",
    )
    p.add_argument(
        "--ab-experiment",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="是否启用A/B生成策略对比评测。",
    )
    p.add_argument(
        "--hit-rate-dashboard-json",
        default="build/Hit_Rate_Dashboard.json",
        help="命中率看板JSON输出路径。",
    )
    p.add_argument(
        "--hit-rate-dashboard-md",
        default="build/Hit_Rate_Dashboard.md",
        help="命中率看板Markdown输出路径。",
    )
    p.add_argument(
        "--enrichment-draft",
        default="build/Auto_KG_Enrichment_Draft.json",
        help="知识盲区反向补图草案输出路径。",
    )
    p.add_argument(
        "--boq-review-queue",
        default="build/BOQ_Manual_Review_Queue.md",
        help="BOQ人工复核队列报告输出路径。",
    )
    p.add_argument(
        "--boq-trust-threshold",
        type=float,
        default=0.78,
        help="BOQ可信度阈值（0-1）。",
    )
    p.add_argument(
        "--boq-hard-gate",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="是否在BOQ可信度不足时直接拦截运行。",
    )
    p.add_argument(
        "--benchmark-auto-expand-min",
        type=int,
        default=120,
        help="评测集最小案例数；不足时自动从图谱扩容。",
    )
    p.add_argument(
        "--benchmark-auto-expand-max",
        type=int,
        default=360,
        help="评测集自动扩容后的最大案例数。",
    )
    p.add_argument(
        "--enforce-auto-generated-lifecycle-gate",
        action=argparse.BooleanOptionalAction,
        default=False,
        help="是否启用自动补丁生命周期门禁（过期/未复核拦截）。",
    )
    p.add_argument(
        "--auto-generated-max-age-days",
        type=int,
        default=120,
        help="自动补丁节点最大允许年龄（天）。",
    )
    return p


async def _run(args: argparse.Namespace) -> int:
    tender_paths = [str(Path(p).expanduser().resolve()) for p in (args.tender or [])]
    boq_path = Path(args.boq).expanduser().resolve()
    kg_root = Path(args.kg_root).expanduser().resolve()
    kg_db = Path(args.kg_db).expanduser().resolve()
    output_path = Path(args.out).expanduser().resolve()
    report_path = Path(args.missing_report).expanduser().resolve()
    docx_out_path = Path(args.docx_out).expanduser().resolve()
    boq_review_queue_path = Path(args.boq_review_queue).expanduser().resolve()
    retrieval_remediation_path = Path(args.retrieval_remediation_report).expanduser().resolve()

    for tender in tender_paths:
        if not Path(tender).exists():
            raise FileNotFoundError(f"tender file not found: {tender}")
    if not boq_path.exists():
        raise FileNotFoundError(f"boq file not found: {boq_path}")
    if not kg_root.exists():
        raise FileNotFoundError(f"kg root not found: {kg_root}")

    boq_payload = await _load_boq_payload(boq_path)
    governance = _build_boq_governance(
        boq_payload=boq_payload,
        trust_threshold=max(0.0, min(float(args.boq_trust_threshold), 1.0)),
    )
    boq_review_queue_saved = _write_boq_manual_review_report(
        governance,
        output_path=boq_review_queue_path,
    )
    boq_payload["governance"] = governance
    if bool(args.boq_hard_gate) and not bool(governance.get("trusted")):
        raise ValueError(
            f"BOQ trust gate blocked: score={float(governance.get('overall_trust_score') or 0.0):.4f}, "
            f"threshold={float(governance.get('trust_threshold') or 0.0):.4f}"
        )

    pipeline = MultiAgentDocPipeline(
        kg_db_path=kg_db,
        self_healing_provider=args.self_heal_provider,
        self_healing_model=args.self_heal_model,
        prefer_human_verified_hits=bool(args.prefer_human_verified_hits),
        min_source_weight=int(args.min_source_weight),
    )

    benchmark_dataset = ensure_benchmark_dataset(
        dataset_path=Path(args.benchmark_dataset).expanduser().resolve(),
        kg_root=kg_root,
        min_cases=max(0, int(args.benchmark_auto_expand_min)),
        max_cases=max(12, int(args.benchmark_auto_expand_max)),
        output_path=Path("build/kg_retrieval_benchmark.auto.json").expanduser().resolve(),
    )

    result = await pipeline.run(
        tender_paths=tender_paths,
        boq_payload=boq_payload,
        graph_root=kg_root,
        output_path=output_path,
        missing_report_path=report_path,
        enable_self_healing=bool(args.self_heal),
        enable_docx_export=bool(args.docx_export),
        docx_output_path=docx_out_path,
        enable_standard_auto_update=bool(args.standard_auto_update),
        run_retrieval_benchmark_gate=bool(args.retrieval_benchmark_gate),
        benchmark_dataset_path=Path(str(benchmark_dataset.get("dataset_path") or args.benchmark_dataset)).expanduser().resolve(),
        benchmark_min_pass_rate=float(args.benchmark_min_pass_rate),
        benchmark_min_avg_mrr=float(args.benchmark_min_avg_mrr),
        benchmark_min_domain_pass_rate=float(args.benchmark_min_domain_pass_rate),
        benchmark_domain_min_cases=max(1, int(args.benchmark_domain_min_cases)),
        enforce_retrieval_gate=bool(args.enforce_retrieval_gate),
        enforce_benchmark_domain_gate=bool(args.enforce_benchmark_domain_gate),
        enable_retrieval_weight_training=bool(args.retrieval_weight_training),
        retrieval_weight_profile_path=Path(args.retrieval_weight_profile).expanduser().resolve(),
        enable_feedback_learning=bool(args.feedback_learning),
        enable_feedback_writeback=bool(args.feedback_writeback),
        region_context=args.region_context,
        bid_date=args.bid_date,
        allow_superseded=bool(args.allow_superseded),
        enforce_standard_reference_gate=True,
        enforce_auto_generated_lifecycle_gate=bool(args.enforce_auto_generated_lifecycle_gate),
        auto_generated_max_age_days=max(7, int(args.auto_generated_max_age_days)),
        regional_plugin_dir=(
            Path(args.regional_plugin_dir).expanduser().resolve() if args.regional_plugin_dir else None
        ),
        create_release_freeze=bool(args.release_freeze),
        release_approver=str(args.release_approver),
        release_signature=f"{args.release_approver}-auto-sign",
        enable_ab_experiment=bool(args.ab_experiment),
        hit_rate_dashboard_json_path=Path(args.hit_rate_dashboard_json).expanduser().resolve(),
        hit_rate_dashboard_md_path=Path(args.hit_rate_dashboard_md).expanduser().resolve(),
        enrichment_draft_path=Path(args.enrichment_draft).expanduser().resolve(),
        boq_governance=governance,
    )
    benchmark = result.get("retrieval_benchmark") if isinstance(result.get("retrieval_benchmark"), dict) else {}
    domain_warnings = (
        result.get("retrieval_benchmark_domain_warnings")
        if isinstance(result.get("retrieval_benchmark_domain_warnings"), list)
        else []
    )
    domain_quality_warnings = (
        result.get("retrieval_benchmark_domain_quality_warnings")
        if isinstance(result.get("retrieval_benchmark_domain_quality_warnings"), list)
        else []
    )
    retrieval_remediation_plan = _build_retrieval_remediation_plan(
        benchmark=benchmark,
        domain_warnings=domain_warnings,
        quality_warnings=domain_quality_warnings,
        min_domain_pass_rate=float(args.benchmark_min_domain_pass_rate),
    )
    retrieval_remediation_saved = _write_retrieval_remediation_report(
        retrieval_remediation_plan,
        output_path=retrieval_remediation_path,
    )
    if output_path.exists():
        try:
            persisted = json.loads(output_path.read_text(encoding="utf-8"))
        except Exception:
            persisted = {}
        if isinstance(persisted, dict):
            persisted["boq_ingestion"] = {
                "requested_path": str(boq_path),
                "source_files": boq_payload.get("source_files") or [],
                "parse_errors": boq_payload.get("parse_errors") or [],
                "stats": boq_payload.get("stats") or {},
                "governance": governance,
                "manual_review_queue_report": boq_review_queue_saved,
                "benchmark_dataset": benchmark_dataset,
                "retrieval_remediation_report": retrieval_remediation_saved,
                "retrieval_remediation_plan": retrieval_remediation_plan,
            }
            output_path.write_text(json.dumps(persisted, ensure_ascii=False, indent=2), encoding="utf-8")

    audit_agent = (result.get("agents") or {}).get("audit_agent") or {}
    score_audit = audit_agent.get("result") or {}
    graph_audit = audit_agent.get("graph_support") or {}
    gaps = result.get("knowledge_gaps") or []

    print("=== Real Project Penetration Run Completed ===")
    print(f"Tender: {', '.join(tender_paths)}")
    print(f"BOQ: {boq_path}")
    boq_stats = boq_payload.get("stats") or {}
    print(
        "BOQ Parse: "
        f"source_files={int(boq_stats.get('source_file_count') or 0)}, "
        f"parsed={int(boq_stats.get('parsed_file_count') or 0)}, "
        f"failed={int(boq_stats.get('failed_file_count') or 0)}, "
        f"items={int(boq_stats.get('item_count') or 0)}"
    )
    boq_governance = boq_payload.get("governance") if isinstance(boq_payload.get("governance"), dict) else {}
    if boq_governance:
        print(
            "BOQ Governance: "
            f"trusted={bool(boq_governance.get('trusted'))}, "
            f"score={float(boq_governance.get('overall_trust_score') or 0.0):.4f}, "
            f"review_items={int(boq_governance.get('manual_review_total') or 0)}, "
            f"queue={boq_review_queue_saved}"
        )
    print(f"Diagnosis JSON: {result.get('saved_at')}")
    print(f"Missing_Knowledge_Report: {result.get('missing_knowledge_report')}")
    print(
        "Hit-Rate Dashboard: "
        f"{(result.get('hit_rate_dashboard_saved') or {}).get('json_path')} | "
        f"{(result.get('hit_rate_dashboard_saved') or {}).get('md_path')}"
    )
    print(f"Auto Enrichment Draft: {(result.get('auto_enrichment_draft') or {}).get('saved_at')}")
    print(f"Production DOCX: {result.get('docx_output')}")
    print(f"Strict Fail-Fast Intercepted: {result.get('intercepted')}")
    print(f"Score Coverage OK: {bool(score_audit.get('ok'))}")
    print(f"Graph Support OK: {bool(graph_audit.get('ok'))}")
    print(f"Knowledge Gaps: {len(gaps)}")
    if benchmark.get("triggered"):
        print(
            "Retrieval Benchmark: "
            f"ok={bool(benchmark.get('ok'))}, "
            f"pass_rate={float(benchmark.get('pass_rate') or 0.0):.4f}, "
            f"avg_mrr={float(benchmark.get('avg_mrr') or 0.0):.4f}"
        )
        if isinstance(domain_warnings, list) and domain_warnings:
            print(f"Retrieval Domain Warnings: {len(domain_warnings)}")
        if isinstance(domain_quality_warnings, list) and domain_quality_warnings:
            print(f"Retrieval Domain Label Warnings: {len(domain_quality_warnings)}")
        print(f"Retrieval Remediation Report: {retrieval_remediation_saved}")
        print(
            "Retrieval Benchmark Dataset: "
            f"cases={int(benchmark_dataset.get('cases_total') or 0)}, "
            f"expanded={bool(benchmark_dataset.get('expanded'))}, "
            f"path={benchmark_dataset.get('dataset_path')}"
        )
    retrieval_weight = (
        result.get("retrieval_weight_profile") if isinstance(result.get("retrieval_weight_profile"), dict) else {}
    )
    if retrieval_weight.get("triggered"):
        print(
            "Retrieval Weight Profile: "
            f"ok={bool(retrieval_weight.get('ok'))}, "
            f"saved_at={retrieval_weight.get('saved_at')}"
        )
    std_upd = result.get("standard_auto_update") if isinstance(result.get("standard_auto_update"), dict) else {}
    if std_upd.get("triggered"):
        print(
            "Standard Auto-Update: "
            f"files_changed={int(std_upd.get('files_changed') or 0)}, "
            f"nodes_updated={int(std_upd.get('nodes_updated') or 0)}"
        )
    feedback = result.get("feedback_learning") if isinstance(result.get("feedback_learning"), dict) else {}
    if feedback.get("triggered"):
        writeback = feedback.get("writeback") if isinstance(feedback.get("writeback"), dict) else {}
        print(
            "Feedback Learning: "
            f"projects_total={int(feedback.get('projects_total') or 0)}, "
            f"node_updates={int(feedback.get('node_updates') or 0)}, "
            f"writeback_ok={bool(writeback.get('ok'))}, "
            f"writeback_nodes={int(writeback.get('nodes_updated') or 0)}"
        )
    chapter_plan = result.get("chapter_response_plan") if isinstance(result.get("chapter_response_plan"), dict) else {}
    if chapter_plan:
        print(
            "Chapter Plan: "
            f"ok={bool(chapter_plan.get('ok'))}, "
            f"chapters={int(chapter_plan.get('chapter_count') or 0)}"
        )
    tactical = result.get("tactical_effects") if isinstance(result.get("tactical_effects"), dict) else {}
    if tactical:
        print(
            "Tactical Effects: "
            f"shield={int(tactical.get('shield_triggered_count') or 0)}, "
            f"booster={int(tactical.get('booster_triggered_count') or 0)}, "
            f"estimated_gain={float(tactical.get('estimated_score_gain') or 0.0):.2f}"
        )
    ab_report = result.get("ab_experiment") if isinstance(result.get("ab_experiment"), dict) else {}
    if ab_report.get("enabled"):
        print(
            "A/B Experiment: "
            f"winner={ab_report.get('winner')}, "
            f"delta={float(ab_report.get('delta') or 0.0):.4f}"
        )
    release = result.get("release_snapshot") if isinstance(result.get("release_snapshot"), dict) else {}
    if release.get("triggered"):
        print(f"Release Snapshot: {release.get('release_id')} | {release.get('release_dir')}")
    release_strategy = result.get("release_strategy") if isinstance(result.get("release_strategy"), dict) else {}
    if release_strategy:
        print(
            "Release Strategy: "
            f"{release_strategy.get('strategy')} "
            f"(reason={release_strategy.get('reason')}, canary={release_strategy.get('canary_ratio')})"
        )
    evidence_stats = result.get("sentence_evidence_stats") or {}
    if evidence_stats:
        print(
            "Sentence Trace: "
            f"total={int(evidence_stats.get('total_sentences') or 0)}, "
            f"traceable={int(evidence_stats.get('traceable_sentences') or 0)}, "
            f"coverage={float(evidence_stats.get('trace_coverage_ratio') or 0.0):.4f}"
        )
    self_heal = result.get("self_healing") or {}
    if self_heal.get("triggered"):
        print(
            "Self-Healing: "
            f"triggered=True, provider={self_heal.get('llm_provider')}, model={self_heal.get('llm_model')}, "
            f"patch_nodes={self_heal.get('patch_nodes')}, used_fallback={self_heal.get('used_fallback')}"
        )
    else:
        print("Self-Healing: triggered=False")

    for i, gap in enumerate(gaps[:20], start=1):
        gtype = str(gap.get("type") or "")
        dim = str(gap.get("dimension") or "")
        kw = "、".join([str(x) for x in (gap.get("required_keywords") or [])[:6]])
        q = str(gap.get("query") or "")
        print(f"GAP[{i}] {gtype} | {dim} | {kw} | query={q}")

    parse_errors = boq_payload.get("parse_errors") or []
    for i, err in enumerate(parse_errors[:10], start=1):
        print(f"BOQ_PARSE_ERROR[{i}] {err.get('file')} | {err.get('error')}")

    return 0


def main() -> int:
    parser = _arg_parser()
    args = parser.parse_args()
    try:
        return asyncio.run(_run(args))
    except Exception as exc:
        print(f"[ERROR] {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
