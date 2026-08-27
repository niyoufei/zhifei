from __future__ import annotations

import asyncio
import tempfile
import threading
from typing import Any, Dict, List

from fastapi import APIRouter, UploadFile, File, HTTPException, Header, BackgroundTasks
from pydantic import BaseModel

from backend.zhifei_autoplan.parsers.tender_parser import TenderParser
from backend.zhifei_autoplan.parsers.boq_parser import BoQParser
from backend.zhifei_autoplan.kg_store import save_kg_bytes, list_kg, set_active_kg, get_active_kg
from backend.zhifei_autoplan.kg_runtime import search_kg
from backend.zhifei_autoplan.tender_store import save_tender_matrix, load_tender_matrix
from backend.zhifei_autoplan.boq_store import load_boq_data, save_boq_data
from backend.auth_store import get_user_by_id, update_balance, log_charge, count_user_actions_since
import jwt
import os
from backend.zhifei_autoplan.orchestrator import (
    new_provider_admission_run_coordinator,
    probe_provider_candidate,
    run_autoplan,
)
from backend.zhifei_autoplan.provider_runtime import (
    ProviderRoutingConfigurationError,
    apply_server_provider_routing,
    build_server_provider_admission_candidates,
    server_provider_admission_required_roles,
)
from backend.zhifei_autoplan.model_reliability import classify_provider_error
from pathlib import Path
import json
from backend.zhifei_autoplan.exporter import export_autoplan_docx_from_file, export_autoplan_docx, export_autoplan_compare_docx
from fastapi.responses import FileResponse
from backend.zhifei_autoplan.job_store import (
    JobLeaseLostError,
    acquire_job_lease,
    create_job,
    get_job,
    heartbeat_job,
    job_lease_active,
    list_jobs,
    cleanup_jobs,
    run_with_job_lease,
    transition_job,
)
from backend.zhifei_autoplan.local_job_queue import submit_isolated_job
from backend.zhifei_autoplan.runtime_events import append_runtime_event
from backend.zhifei_autoplan.output_artifacts import sanitize_output_payload
from backend.zhifei_autoplan.plan_store import save_plan, load_plan
from backend.zhifei_autoplan.optimizer import optimize_sections
from backend.zhifei_autoplan.variant_cycle import reserve_variant_ids

try:
    from backend.zhifei_autoplan.local_adapter_shim import (
        block_export_response as _local_adapter_block_export_response,
        map_output as _local_adapter_map_output,
    )
except Exception:
    _local_adapter_block_export_response = None
    _local_adapter_map_output = None

router = APIRouter(prefix="/autoplan", tags=["Zhifei AutoPlan"])
JWT_SECRET = os.environ.get("ZF_JWT_SECRET", "change-me")
JWT_ALG = "HS256"
COST_PER_JOB = int(os.environ.get("ZF_JOB_COST", "1"))


def _require_generation_sources(payload: Dict[str, Any], project_id: str | None) -> None:
    if bool(payload.get("dry_run")):
        return
    if not str(project_id or "").strip():
        raise HTTPException(
            status_code=409,
            detail={
                "code": "PROJECT_SCOPE_REQUIRED",
                "message": "真实生成必须绑定明确项目，禁止使用全局资料池。",
            },
        )
    missing: list[str] = []
    tender_source = load_tender_matrix(project_id=project_id) or {}
    boq_source = load_boq_data(project_id=project_id) or {}
    if not isinstance(tender_source.get("outline"), list) or not tender_source.get("outline"):
        missing.append("tender")
    if not isinstance(boq_source.get("items"), list) or not boq_source.get("items"):
        missing.append("boq")
    if missing:
        raise HTTPException(
            status_code=409,
            detail={
                "code": "MANDATORY_SOURCE_NOT_READY",
                "message": "招标/答疑与工程量清单必须全部解析成功后才能生成。",
                "missing": missing,
            },
        )


def _route_generation_or_503(payload: Dict[str, Any]) -> Dict[str, Any]:
    try:
        routed = apply_server_provider_routing(payload)
        # This compatibility route exports deterministic DOCX directly and
        # never calls the professional document-render model.
        routed["_provider_admission_required_roles"] = [
            role
            for role in (routed.get("_provider_admission_required_roles") or [])
            if str(role or "").strip().lower() != "document_render"
        ]
        return routed
    except ProviderRoutingConfigurationError as exc:
        raise HTTPException(
            status_code=503,
            detail={
                "code": "MODEL_PROVIDER_CONFIGURATION_BLOCKED",
                "message": "服务端模型路由未配置完整或尚不支持准入，已阻止生成。",
            },
        ) from exc


async def _admit_server_provider_chain(payload: Dict[str, Any]) -> tuple[Any, List[Dict[str, Any]]]:
    """Freshly admit the server-owned chain and return ephemeral text credentials.

    The returned credentials live only in memory.  Callers must never place the
    returned chain in a job record, result, checkpoint, audit record or response.
    """

    coordinator = new_provider_admission_run_coordinator(payload)
    candidates = build_server_provider_admission_candidates()
    required_roles = server_provider_admission_required_roles(
        candidates,
        require_document_render=False,
    )
    snapshot = await coordinator.admit_chain_once(
        candidates=candidates,
        probe=probe_provider_candidate,
        required_roles=required_roles,
    )
    if not bool(snapshot.get("generation_allowed")):
        raise HTTPException(
            status_code=503,
            detail={
                "code": "MODEL_PROVIDER_ADMISSION_BLOCKED",
                "message": "模型供应商准入未通过，已在调用模型前停止。",
                "action": "请在供应商准入页面检查凭据、模型、配额、流式能力和文档渲染槽位。",
            },
        )
    admitted_identities = {
        str(row.get("identity_digest") or "")
        for row in (snapshot.get("admitted_chain") or [])
        if isinstance(row, dict) and str(row.get("identity_digest") or "")
    }
    chain: List[Dict[str, Any]] = []
    for candidate in coordinator.bound_candidates:
        if not str(candidate.role or "").startswith("text_"):
            continue
        if candidate.identity_digest not in admitted_identities:
            continue
        chain.append(
            {
                "slot": candidate.slot,
                "role": candidate.role,
                "provider": candidate.provider,
                "model": candidate.model,
                "api_key": candidate.credential,
            }
        )
    if not chain:
        raise HTTPException(
            status_code=503,
            detail={
                "code": "MODEL_PROVIDER_ADMISSION_EMPTY",
                "message": "没有可用的已准入文本模型，已停止调用。",
            },
        )
    return coordinator, chain


def _stable_job_error(
    exc: BaseException,
    *,
    provider: str = "",
    model: str = "",
) -> str:
    """Project arbitrary exceptions into a stable, Chinese, credential-free error."""

    raw = str(exc or "")
    try:
        parsed = json.loads(raw)
    except Exception:
        parsed = None
    if isinstance(parsed, dict):
        code = str(parsed.get("code") or "").strip()
        if code and all(ch.isupper() or ch.isdigit() or ch == "_" for ch in code):
            return json.dumps(
                {
                    "code": code[:80],
                    "message": str(parsed.get("message") or "任务执行失败，系统已安全停止。")[:300],
                    "action": str(parsed.get("action") or "请根据错误码检查资料和模型准入状态后重试。")[:300],
                },
                ensure_ascii=False,
            )
    provider_error = classify_provider_error(
        exc,
        provider=str(provider or ""),
        model=str(model or ""),
    )
    return json.dumps(
        {
            "code": str(provider_error.get("code") or "AUTOPLAN_RUNTIME_FAILED")[:80],
            "message": str(provider_error.get("user_message") or "任务执行失败，系统已安全停止。")[:300],
            "action": str(provider_error.get("action") or "请检查资料和供应商准入状态后重试。")[:300],
        },
        ensure_ascii=False,
    )


def _local_adapter_issue(code: str, message: str, *, variant_index: int | None = None) -> Dict[str, Any]:
    issue: Dict[str, Any] = {"code": code, "message": message, "severity": "error"}
    if variant_index is not None:
        issue["variant_index"] = variant_index
    return issue


def _local_adapter_block(issues: List[Dict[str, Any]]) -> Dict[str, Any]:
    if _local_adapter_block_export_response is not None:
        return _local_adapter_block_export_response(issues)
    return {"ok": False, "status": "blocked", "export_allowed": False, "issues": issues}


def _local_adapter_gate_results(results: List[Dict[str, Any]]) -> Dict[str, Any]:
    if _local_adapter_map_output is None:
        return {
            "export_allowed": False,
            "results": results,
            "issues": [_local_adapter_issue("ADAPTER_IMPORT_FAILURE", "local adapter shim import failed")],
        }
    issues: List[Dict[str, Any]] = []
    gated_results: List[Dict[str, Any]] = []
    for idx, result in enumerate(results):
        if not isinstance(result, dict):
            issues.append(_local_adapter_issue("ADAPTER_OUTPUT_INVALID", "variant output is not a dict", variant_index=idx))
            continue
        try:
            envelope = _local_adapter_map_output(result)
        except Exception:
            issues.append(
                _local_adapter_issue(
                    "ADAPTER_HOOK_FAILURE",
                    "生成结果适配校验失败，已阻止导出。",
                    variant_index=idx,
                )
            )
            continue
        adapter_view = {
            "status": envelope.get("status"),
            "export_allowed": bool(envelope.get("export_allowed")),
            "issues": envelope.get("issues") or [],
            "hard_gates": envelope.get("hard_gates") or [],
            "evidence_summary": envelope.get("evidence_summary") or {},
        }
        result["local_adapter"] = adapter_view
        if not adapter_view["export_allowed"]:
            issues.extend(
                issue if isinstance(issue, dict) else _local_adapter_issue("LOCAL_ADAPTER_EXPORT_BLOCKED", str(issue), variant_index=idx)
                for issue in adapter_view["issues"]
            )
        gated_results.append(result)
    return {"export_allowed": not issues, "results": gated_results, "issues": issues}


def _local_adapter_job_error(issues: List[Dict[str, Any]]) -> str:
    return json.dumps(_local_adapter_block(issues), ensure_ascii=False)


def _load_field_alias() -> dict:
    try:
        from pathlib import Path
        import json
        cfg_path = Path("backend/data/autoplan/config.json")
        if cfg_path.exists():
            cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
            alias = cfg.get("job_list_field_alias")
            if isinstance(alias, dict):
                return {str(k): {str(s) for s in v} for k, v in alias.items() if isinstance(v, list)}
    except Exception:
        pass
    return {}


def _parse_fields(fields: list[str] | None) -> set[str]:
    if not fields:
        return set()
    parsed: set[str] = set()
    for f in fields:
        if not isinstance(f, str):
            continue
        for part in f.split(","):
            p = part.strip()
            if p:
                parsed.add(p)
    # alias expand (configurable)
    alias = _load_field_alias()
    if not alias:
        alias = {
            "time": {"created_at", "updated_at"},
            "meta": {"job_id", "status", "created_at", "updated_at"},
            "result_min": {"job_id", "status", "result"},
        }
    expanded: set[str] = set()
    for p in parsed:
        if p in alias:
            expanded.update(alias[p])
        else:
            expanded.add(p)
    return expanded


def _job_list_default_fields() -> set[str]:
    env_fields = os.environ.get("ZF_JOB_LIST_FIELDS")
    if env_fields:
        return {s.strip() for s in env_fields.split(",") if s.strip()}
    try:
        from pathlib import Path
        import json
        cfg_path = Path("backend/data/autoplan/config.json")
        if cfg_path.exists():
            cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
            fields = cfg.get("job_list_default_fields")
            if isinstance(fields, list) and fields:
                return {str(s) for s in fields}
    except Exception:
        pass
    return {"job_id", "status", "created_at", "updated_at", "result", "error"}


def _job_list_field_alias() -> dict:
    alias = _load_field_alias()
    if alias:
        return alias
    return {
        "time": {"created_at", "updated_at"},
        "meta": {"job_id", "status", "created_at", "updated_at"},
        "result_min": {"job_id", "status", "result"},
    }


def _auth_user(authorization: str | None):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="missing token")
    token = authorization.split(" ", 1)[1].strip()
    try:
        data = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        user_id = int(data.get("sub"))
        user = get_user_by_id(user_id)
        if not user:
            raise HTTPException(status_code=401, detail="invalid user")
        return user
    except Exception:
        raise HTTPException(status_code=401, detail="invalid token")


def _is_admin(authorization: str | None) -> bool:
    admin_key = os.environ.get("ZF_ADMIN_KEY", "")
    return bool(admin_key and authorization == f"Bearer {admin_key}")


def _charge(user: dict, cost: int, action: str):
    if user.get("balance", 0) < cost:
        raise HTTPException(status_code=402, detail="insufficient balance")
    used = count_user_actions_since(user["id"], 24 * 3600)
    limit = int(user.get("daily_limit") or os.environ.get("ZF_DAILY_LIMIT", "50"))
    if used >= limit:
        raise HTTPException(status_code=429, detail="daily limit exceeded")
    update_balance(user["id"], -cost)
    log_charge(user["id"], action, cost)


def _audit(action: str, user_id: int | None = None, detail: dict | None = None):
    try:
        from datetime import datetime
        path = Path("backend/data/audit/autoplan.jsonl")
        path.parent.mkdir(parents=True, exist_ok=True)
        rec = {
            "ts": datetime.utcnow().isoformat(),
            "action": action,
            "user_id": user_id,
            "detail": detail or {},
        }
        with path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    except Exception:
        pass


def _read_audit(
    limit: int = 200,
    action: str | None = None,
    user_id: int | None = None,
    ts_after: str | None = None,
    ts_before: str | None = None,
):
    path = Path("backend/data/audit/autoplan.jsonl")
    if not path.exists():
        return []
    lines = path.read_text(encoding="utf-8").splitlines()
    items = []
    for line in reversed(lines):
        if not line.strip():
            continue
        try:
            rec = json.loads(line)
        except Exception:
            continue
        if action and rec.get("action") != action:
            continue
        if user_id is not None and rec.get("user_id") != user_id:
            continue
        if ts_after and str(rec.get("ts") or "") < ts_after:
            continue
        if ts_before and str(rec.get("ts") or "") > ts_before:
            continue
        items.append(rec)
        if len(items) >= limit:
            break
    return items


def _audit_summary(limit: int = 10000, user_id: int | None = None, by_user: bool = False):
    path = Path("backend/data/audit/autoplan.jsonl")
    if not path.exists():
        return {"by_action": {}, "by_user": {}}
    lines = path.read_text(encoding="utf-8").splitlines()
    by_action = {}
    by_user_count = {}
    n = 0
    for line in reversed(lines):
        if n >= limit:
            break
        if not line.strip():
            continue
        try:
            rec = json.loads(line)
        except Exception:
            continue
        if user_id is not None and rec.get("user_id") != user_id:
            continue
        n += 1
        act = rec.get("action") or "unknown"
        by_action[act] = by_action.get(act, 0) + 1
        if by_user:
            uid = rec.get("user_id")
            by_user_count[uid] = by_user_count.get(uid, 0) + 1
    return {"by_action": by_action, "by_user": by_user_count, "total": n}


def _audit_stats_by_day(limit_days: int = 30, user_id: int | None = None):
    path = Path("backend/data/audit/autoplan.jsonl")
    if not path.exists():
        return {"by_day": {}, "total": 0}
    lines = path.read_text(encoding="utf-8").splitlines()
    by_day = {}
    total = 0
    try:
        from datetime import datetime, timedelta
        cutoff = (datetime.utcnow() - timedelta(days=limit_days)).isoformat()[:10]
    except Exception:
        cutoff = ""
    for line in lines:
        if not line.strip():
            continue
        try:
            rec = json.loads(line)
        except Exception:
            continue
        if user_id is not None and rec.get("user_id") != user_id:
            continue
        ts = (rec.get("ts") or "")[:10]
        if cutoff and ts < cutoff:
            continue
        by_day[ts] = by_day.get(ts, 0) + 1
        total += 1
    return {"by_day": dict(sorted(by_day.items())), "total": total}


async def _save_upload(uf: UploadFile) -> str:
    # 临时落地到磁盘，供 pdfplumber/pandas 读取
    data = await uf.read()
    if not data:
        raise HTTPException(status_code=400, detail=f"空文件：{uf.filename}")
    suffix = f"_{uf.filename}"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as f:
        f.write(data)
        return f.name


@router.post("/tender/parse")
async def parse_tender(
    files: List[UploadFile] = File(...),
    project_id: str | None = None,
    authorization: str | None = Header(default=None),
):
    """
    Module 1：招标文件指数级解析
    - MECE：输出 6 大维度指标，互不包含、完全穷举
    """
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "tender_parse")
    if not files:
        raise HTTPException(status_code=400, detail="未上传文件")
    if not str(project_id or "").strip():
        raise HTTPException(status_code=422, detail={"code": "PROJECT_SCOPE_REQUIRED"})
    paths = await asyncio.gather(*[_save_upload(f) for f in files])
    parser = TenderParser()
    matrix = await parser.parse(paths)
    matrix_payload = matrix.model_dump()
    if not isinstance(matrix_payload.get("outline"), list) or not matrix_payload.get("outline"):
        raise HTTPException(status_code=422, detail={"code": "TENDER_PARSE_NOT_READY"})
    matrix_payload["parse_status"] = "ready"
    matrix_payload["project_id"] = str(project_id).strip()
    saved_at = save_tender_matrix(matrix_payload, project_id=project_id)
    return {"matrix": matrix_payload, "saved_at": saved_at}


@router.post("/boq/parse")
async def parse_boq(
    file: UploadFile = File(...),
    project_id: str | None = None,
    authorization: str | None = Header(default=None),
):
    """
    Module 2：清单智能识别与统计
    - MECE：清单->工序->资源 三段映射
    """
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "boq_parse")
    if not file:
        raise HTTPException(status_code=400, detail="未上传文件")
    if not str(project_id or "").strip():
        raise HTTPException(status_code=422, detail={"code": "PROJECT_SCOPE_REQUIRED"})
    path = await _save_upload(file)
    parser = BoQParser()
    items, stats = await parser.parse(path)
    if not items:
        raise HTTPException(status_code=422, detail={"code": "BOQ_PARSE_NOT_READY"})
    payload = {
        "items": [it.model_dump() for it in items],
        "stats": stats,
        "parse_status": "ready",
        "project_id": str(project_id).strip(),
    }
    saved_at = save_boq_data(payload, project_id=project_id)
    return {**payload, "saved_at": saved_at}


@router.post("/plan/save")
async def save_plan_api(req: PlanRequest, project_id: str | None = None, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    path = save_plan(req.model_dump(), project_id=project_id)
    _audit("plan_save", user_id=user["id"], detail={"path": path, "project_id": project_id})
    return {"ok": True, "saved_at": path}


@router.get("/plan/get")
async def get_plan_api(project_id: str | None = None, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _audit("plan_get", user_id=user["id"], detail={"project_id": project_id})
    return {"ok": True, "plan": load_plan(project_id=project_id) or {}}


class ActivateKGRequest(BaseModel):
    kg_id: str


class GenerateRequest(BaseModel):
    project_id: str | None = None
    topic: str
    outline: List[str] = []
    requirements: List[str] = []
    chapter_requirements: dict | None = None
    # Reusable project context is accepted as separate fields so the
    # orchestrator can place it before the dynamic chapter instruction.
    chapter_summaries: List[dict | str] = []
    project_stage_context: str | None = None
    common_construction_requirements: List[str] = []
    provider: str | None = None
    model: str | None = None
    provider_chain: List[dict] | None = None
    providers: List[str] = []
    model_map: dict | None = None
    style: dict | None = None
    variants: int = 1
    chapter_pages: dict | None = None
    quality_strict: bool | None = None
    auto_remediate: bool = True
    remediate_mode: str = "template"  # template | llm
    compare_mode: str = "full"  # full | summary
    compare_max_chars: int = 800
    compare_titles: list[str] | None = None
    api_key: str | None = None
    base_url: str | None = None
    secret_key: str | None = None
    token_url: str | None = None
    dry_run: bool = False
    generate_images: bool = True


class PlanRequest(BaseModel):
    outline: List[str]
    style: dict = {}
    variants: int = 1
    chapter_requirements: dict = {}
    chapter_pages: dict = {}
    quality_strict: bool = True
    auto_remediate: bool = True
    remediate_mode: str = "template"
    compare_mode: str = "full"
    compare_max_chars: int = 800
    compare_titles: list[str] | None = None


class OptimizeRequest(BaseModel):
    titles: List[str]
    instruction: str = "请在保持证据引用的前提下优化本章表达。"
    provider: str | None = None
    model: str | None = None
    api_key: str | None = None
    base_url: str | None = None
    secret_key: str | None = None
    token_url: str | None = None


class JobListFilterRequest(BaseModel):
    job_ids: list[str] | None = None
    limit: int = 200
    status: str | None = None
    has_compare: bool | None = None
    updated_after: float | None = None
    updated_before: float | None = None
    recent_hours: int | None = None
    created_after: float | None = None
    created_before: float | None = None
    recent_days: int | None = None
    completed_after: float | None = None
    completed_before: float | None = None
    completed_recent_hours: int | None = None
    fields: list[str] | None = None
    full: bool = False


class AuditQueryRequest(BaseModel):
    limit: int = 200
    action: str | None = None
    user_id: int | None = None
    ts_after: str | None = None
    ts_before: str | None = None


class AuditExportRequest(BaseModel):
    fmt: str = "json"  # json|csv|xlsx
    limit: int = 1000
    action: str | None = None
    user_id: int | None = None
    ts_after: str | None = None
    ts_before: str | None = None
    filename: str | None = None


class AuditExportListRequest(BaseModel):
    user_id: int | None = None
    limit: int = 100


class AuditExportCleanupRequest(BaseModel):
    user_id: int | None = None
    older_than_days: int | None = None
    keep_latest_n: int | None = None


def run_legacy_generation_job(job_id: str, payload: dict) -> None:
    """Spawn-safe compatibility worker shared by single and batch routes."""

    heartbeat_stop = threading.Event()
    heartbeat_thread: threading.Thread | None = None

    lease_record = acquire_job_lease(job_id)
    if lease_record is None:
        return
    lease_attempt_id = str(lease_record.get("attempt_id") or "")
    lease_owner_instance_id = str(lease_record.get("owner_instance_id") or "")
    if not lease_attempt_id or not lease_owner_instance_id:
        raise RuntimeError("job_lease_acquisition_invalid")

    def _lease_active() -> bool:
        return job_lease_active(
            job_id,
            attempt_id=lease_attempt_id,
            owner_instance_id=lease_owner_instance_id,
        )

    def _status() -> str:
        return str((get_job(job_id) or {}).get("status") or "").strip().lower()

    def _cancel_requested() -> bool:
        return _status() in {"cancel_requested", "cancelled"} or not _lease_active()

    def _lease_side_effect(callback: Any, *args: Any, **kwargs: Any) -> Any:
        return run_with_job_lease(
            job_id,
            attempt_id=lease_attempt_id,
            owner_instance_id=lease_owner_instance_id,
            callback=callback,
            callback_args=tuple(args),
            callback_kwargs=dict(kwargs),
        )

    def _running_side_effect(callback: Any, *args: Any, **kwargs: Any) -> Any:
        return run_with_job_lease(
            job_id,
            attempt_id=lease_attempt_id,
            owner_instance_id=lease_owner_instance_id,
            allowed_statuses={"running"},
            callback=callback,
            callback_args=tuple(args),
            callback_kwargs=dict(kwargs),
        )

    def _append_active_event(event: str, **fields: Any) -> bool:
        try:
            _running_side_effect(append_runtime_event, job_id, event, **fields)
            return True
        except JobLeaseLostError:
            return False

    def _mark_cancelled() -> None:
        if not _lease_active():
            return
        prior_progress = (get_job(job_id) or {}).get("progress") or {}
        try:
            from backend.zhifei_autoplan.generation_checkpoint import (
                mark_checkpoint_namespace_interrupted,
            )

            scopes = _lease_side_effect(mark_checkpoint_namespace_interrupted, job_id)
            checkpoint_projection = {
                "status": (
                    "interrupted_recoverable" if scopes else "interrupted_empty"
                ),
                "saved_chapter_count": sum(
                    int(item.get("saved_chapter_count") or 0)
                    for item in scopes
                    if isinstance(item, dict)
                ),
                "scopes": scopes,
            }
            seal_failed = False
        except JobLeaseLostError:
            return
        except Exception as seal_error:
            checkpoint_projection = {
                "status": "interruption_seal_failed",
                "saved_chapter_count": 0,
                "error_code": "CHECKPOINT_INTERRUPTION_SEAL_FAILED",
                "error_type": type(seal_error).__name__,
            }
            seal_failed = True
        transitioned = transition_job(
            job_id,
            allowed_from={"running", "cancel_requested"},
            status="cancelled",
            expected_attempt_id=lease_attempt_id,
            expected_owner_instance_id=lease_owner_instance_id,
            revoke_lease=True,
            error={
                "code": (
                    "JOB_CANCELLED_CHECKPOINT_SEAL_FAILED"
                    if seal_failed
                    else "JOB_CANCELLED"
                ),
                "message": (
                    "用户已取消任务，但检查点封存失败；该故障已显式记录。"
                    if seal_failed
                    else "用户已取消任务。"
                ),
                "action": (
                    "先检查检查点存储与权限，再决定是否从先前可信断点恢复。"
                    if seal_failed
                    else "可从已保存的可信检查点显式恢复。"
                ),
            },
            progress={
                "percent": min(99, int(prior_progress.get("percent") or 0)),
                "phase": "generation",
                "work_state": "idle",
                "checkpoint": checkpoint_projection,
            },
        )
        if transitioned is not None:
            append_runtime_event(job_id, "legacy_job_cancelled")

    def _heartbeat() -> None:
        while not heartbeat_stop.wait(5.0):
            updated = heartbeat_job(
                job_id,
                activity="兼容任务正在隔离生成",
                progress_updates={"phase": "generation", "work_state": "processing"},
                expected_attempt_id=lease_attempt_id,
                expected_owner_instance_id=lease_owner_instance_id,
                allowed_statuses={"running"},
            )
            if updated is None:
                heartbeat_stop.set()
                return

    try:
        if _cancel_requested():
            _mark_cancelled()
            return
        started = transition_job(
            job_id,
            allowed_from={"running"},
            status="running",
            expected_attempt_id=lease_attempt_id,
            expected_owner_instance_id=lease_owner_instance_id,
            progress={"phase": "generation", "work_state": "processing", "percent": 1},
        )
        if started is None:
            return
        _append_active_event("legacy_job_started")
        heartbeat_thread = threading.Thread(
            target=_heartbeat,
            name=f"legacy-autoplan-heartbeat-{job_id[:8]}",
            daemon=True,
        )
        heartbeat_thread.start()
        local_payload = json.loads(json.dumps(payload))
        provider_admission_run = (
            new_provider_admission_run_coordinator(local_payload)
            if bool(local_payload.get("_provider_admission_required"))
            else None
        )
        variants = int(local_payload.get("variants") or 1)
        variant_ids = local_payload.get("_variant_ids")
        if not isinstance(variant_ids, list) or not variant_ids:
            variant_ids = _running_side_effect(
                reserve_variant_ids,
                project_id=str(local_payload.get("project_id") or "").strip() or None,
                count=max(1, variants),
                explicit_variant_id=local_payload.get("variant_id"),
                explicit_template_id=local_payload.get("logic_template_id")
                or local_payload.get("logic_template"),
            )
        results = []
        for variant_id in variant_ids:
            if _cancel_requested():
                _mark_cancelled()
                return
            local_payload["variant_id"] = int(variant_id)
            local_payload["_job_id"] = job_id
            local_payload["_checkpoint_namespace"] = job_id
            local_payload["_cancel_callback"] = _cancel_requested
            local_payload["_checkpoint_write_guard"] = _lease_side_effect
            if provider_admission_run is not None:
                local_payload["_provider_admission_run_coordinator"] = provider_admission_run
            result = asyncio.run(run_autoplan(local_payload))
            if _cancel_requested():
                _mark_cancelled()
                return
            results.append(result)
        if _cancel_requested():
            _mark_cancelled()
            return
        gate = _local_adapter_gate_results(results)
        if not gate["export_allowed"]:
            failed = transition_job(
                job_id,
                allowed_from={"running"},
                status="failed",
                expected_attempt_id=lease_attempt_id,
                expected_owner_instance_id=lease_owner_instance_id,
                revoke_lease=True,
                result=_local_adapter_block(gate["issues"]),
                error=_local_adapter_job_error(gate["issues"]),
                progress={"work_state": "idle", "percent": 99},
            )
            if failed is not None:
                append_runtime_event(job_id, "legacy_job_failed", code="LOCAL_ADAPTER_BLOCKED")
            elif _status() == "cancel_requested":
                _mark_cancelled()
            return
        results = gate["results"]

        output_dir = Path("build")
        _running_side_effect(output_dir.mkdir, parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(
            prefix=f"zhifei-autoplan-{job_id}-",
            dir=str(output_dir),
        ) as staging_dir_value:
            staging_dir = Path(staging_dir_value)
            staged_json = staging_dir / f"autoplan_{job_id}.json"
            staged_json.write_text(
                json.dumps(
                    {"variants": sanitize_output_payload(results)},
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            staged_docx: list[Path] = []
            staged_compare: list[Path] = []
            for index, variant in enumerate(results, start=1):
                staged_variant = staging_dir / f"autoplan_{job_id}_v{index}.docx"
                export_autoplan_docx(variant, str(staged_variant))
                staged_docx.append(staged_variant)
                staged_comparison = (
                    staging_dir / f"autoplan_{job_id}_compare_v{index}.docx"
                )
                export_autoplan_compare_docx(variant, str(staged_comparison))
                staged_compare.append(staged_comparison)

            final_json = output_dir / staged_json.name
            final_docx = [output_dir / path.name for path in staged_docx]
            final_compare = [output_dir / path.name for path in staged_compare]

            def _publish_outputs() -> None:
                output_dir.mkdir(parents=True, exist_ok=True)
                staged_json.replace(final_json)
                for staged_path, final_path in zip(staged_docx, final_docx):
                    staged_path.replace(final_path)
                for staged_path, final_path in zip(staged_compare, final_compare):
                    staged_path.replace(final_path)

            # Rendering happens outside the global job-store lock.  Only the
            # short, same-filesystem publish is fenced against cancellation or
            # reconciliation, so a stale worker can never expose its outputs.
            _running_side_effect(_publish_outputs)
            output_files = {
                "json": str(final_json),
                "docx": [str(path) for path in final_docx],
                "compare_docx": [str(path) for path in final_compare],
            }
        transition = transition_job(
            job_id,
            allowed_from={"running"},
            status="succeeded",
            expected_attempt_id=lease_attempt_id,
            expected_owner_instance_id=lease_owner_instance_id,
            revoke_lease=True,
            result=output_files,
            error=None,
            progress={"phase": "done", "work_state": "idle", "percent": 100},
        )
        if transition is not None:
            append_runtime_event(job_id, "legacy_job_succeeded")
        elif _status() == "cancel_requested":
            _mark_cancelled()
    except JobLeaseLostError:
        if _status() == "cancel_requested":
            _mark_cancelled()
        return
    except Exception as exc:
        if _cancel_requested():
            _mark_cancelled()
        else:
            failed = transition_job(
                job_id,
                allowed_from={"running"},
                status="failed",
                expected_attempt_id=lease_attempt_id,
                expected_owner_instance_id=lease_owner_instance_id,
                revoke_lease=True,
                error=_stable_job_error(
                    exc,
                    provider=str(payload.get("provider") or ""),
                    model=str(payload.get("model") or ""),
                ),
                progress={"phase": "generation", "work_state": "idle", "percent": 99},
            )
            if failed is not None:
                append_runtime_event(job_id, "legacy_job_failed", error_type=type(exc).__name__)
            elif _status() == "cancel_requested":
                _mark_cancelled()
    finally:
        heartbeat_stop.set()
        if heartbeat_thread is not None and heartbeat_thread.is_alive():
            heartbeat_thread.join(timeout=0.25)


async def _wait_for_legacy_job(job_id: str, *, timeout_seconds: float = 3600.0) -> dict:
    deadline = asyncio.get_running_loop().time() + max(1.0, float(timeout_seconds))
    while True:
        record = get_job(job_id)
        if not record:
            raise HTTPException(status_code=500, detail={"code": "JOB_RECORD_LOST"})
        status = str(record.get("status") or "").strip().lower()
        if status not in {"queued", "running", "cancel_requested"}:
            return record
        if asyncio.get_running_loop().time() >= deadline:
            raise HTTPException(
                status_code=504,
                detail={"code": "LEGACY_SYNC_WAIT_TIMEOUT", "job_id": job_id},
            )
        await asyncio.sleep(0.1)


@router.post("/kg/upload")
async def upload_kg(file: UploadFile = File(...)):
    """
    上传你已有的知识图谱（JSON），系统只做存档与追溯，不重新构建。
    """
    if not file:
        raise HTTPException(status_code=400, detail="未上传文件")
    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="空文件")
    try:
        # 仅做 JSON 结构校验，避免错误文件
        import json as _json
        _json.loads(data.decode("utf-8", errors="replace"))
    except Exception:
        raise HTTPException(status_code=400, detail="仅支持 JSON 知识图谱文件")
    meta = save_kg_bytes(data, file.filename)
    return {"ok": True, "kg": meta}


@router.get("/kg/list")
async def list_kg_files():
    return {"ok": True, "items": list_kg()}


@router.get("/kg/active")
async def get_active_kg_file():
    return {"ok": True, "active": get_active_kg()}


@router.post("/kg/activate")
async def activate_kg(req: ActivateKGRequest):
    try:
        rec = set_active_kg(req.kg_id)
        return {"ok": True, "active": rec}
    except ValueError:
        raise HTTPException(status_code=404, detail="kg_id not found")


@router.get("/kg/search")
async def search_kg_api(q: str, top_k: int = 6):
    return search_kg(q, top_k=top_k)


@router.post("/generate")
async def generate_plan(req: GenerateRequest, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "autoplan_generate")
    payload = req.model_dump()
    pid = str(payload.get("project_id") or "").strip() or None
    plan = load_plan(project_id=pid) or {}
    tender = load_tender_matrix(project_id=pid) or {}
    if not payload.get("outline"):
        payload["outline"] = plan.get("outline") or []
    if not payload.get("outline"):
        payload["outline"] = tender.get("outline") or []
    if payload.get("chapter_requirements") is None:
        payload["chapter_requirements"] = plan.get("chapter_requirements") or {}
    if not payload.get("chapter_requirements"):
        payload["chapter_requirements"] = tender.get("chapter_requirements") or {}
    if payload.get("style") is None:
        payload["style"] = plan.get("style") or {}
    if not payload.get("style"):
        payload["style"] = tender.get("style") or {}
    if payload.get("chapter_pages") is None:
        payload["chapter_pages"] = plan.get("chapter_pages") or {}
    if not payload.get("chapter_pages"):
        payload["chapter_pages"] = tender.get("chapter_pages") or {}
    if payload.get("quality_strict") is None:
        payload["quality_strict"] = plan.get("quality_strict", True)
    if payload.get("auto_remediate") is None:
        payload["auto_remediate"] = plan.get("auto_remediate", True)
    if payload.get("remediate_mode") is None:
        payload["remediate_mode"] = plan.get("remediate_mode", "template")
    if payload.get("compare_mode") is None:
        payload["compare_mode"] = plan.get("compare_mode", "full")
    if payload.get("compare_max_chars") is None:
        payload["compare_max_chars"] = plan.get("compare_max_chars", 800)
    if payload.get("compare_titles") is None:
        payload["compare_titles"] = plan.get("compare_titles")
    if not payload.get("variants"):
        payload["variants"] = plan.get("variants") or 1

    _require_generation_sources(payload, pid)
    payload = _route_generation_or_503(payload)
    variants = int(payload.get("variants") or 1)
    payload["_variant_ids"] = reserve_variant_ids(
        project_id=pid,
        count=max(1, variants),
        explicit_variant_id=payload.get("variant_id"),
        explicit_template_id=payload.get("logic_template_id") or payload.get("logic_template"),
    )
    job_id = create_job(payload, user_id=user["id"])
    submit_isolated_job(job_id, run_legacy_generation_job, job_id, payload)
    record = await _wait_for_legacy_job(job_id)
    if str(record.get("status") or "").strip().lower() != "succeeded":
        blocked = record.get("result")
        if isinstance(blocked, dict) and blocked.get("ok") is False:
            return blocked
        raise HTTPException(
            status_code=500,
            detail={"code": "LEGACY_GENERATION_FAILED", "job_id": job_id, "error": record.get("error")},
        )
    files = record.get("result") if isinstance(record.get("result"), dict) else {}
    out_path = Path(str(files.get("json") or ""))
    data = json.loads(out_path.read_text(encoding="utf-8"))
    results = data.get("variants") if isinstance(data.get("variants"), list) else []
    docx_files = list(files.get("docx") or [])
    _audit(
        "generate",
        user_id=user["id"],
        detail={"job_id": job_id, "variants": len(results), "docx": docx_files, "project_id": pid},
    )
    return {"ok": True, "saved_at": str(out_path), "docx": docx_files, "result": results}


@router.post("/generate_async")
async def generate_plan_async(req: GenerateRequest, background_tasks: BackgroundTasks, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "autoplan_generate_async")
    payload = req.model_dump()
    pid = str(payload.get("project_id") or "").strip() or None
    plan = load_plan(project_id=pid) or {}
    tender = load_tender_matrix(project_id=pid) or {}
    if not payload.get("outline"):
        payload["outline"] = plan.get("outline") or []
    if not payload.get("outline"):
        payload["outline"] = tender.get("outline") or []
    if payload.get("chapter_requirements") is None:
        payload["chapter_requirements"] = plan.get("chapter_requirements") or {}
    if not payload.get("chapter_requirements"):
        payload["chapter_requirements"] = tender.get("chapter_requirements") or {}
    if payload.get("style") is None:
        payload["style"] = plan.get("style") or {}
    if not payload.get("style"):
        payload["style"] = tender.get("style") or {}
    if payload.get("chapter_pages") is None:
        payload["chapter_pages"] = plan.get("chapter_pages") or {}
    if not payload.get("chapter_pages"):
        payload["chapter_pages"] = tender.get("chapter_pages") or {}
    if payload.get("quality_strict") is None:
        payload["quality_strict"] = plan.get("quality_strict", True)
    if payload.get("auto_remediate") is None:
        payload["auto_remediate"] = plan.get("auto_remediate", True)
    if payload.get("remediate_mode") is None:
        payload["remediate_mode"] = plan.get("remediate_mode", "template")
    if payload.get("compare_mode") is None:
        payload["compare_mode"] = plan.get("compare_mode", "full")
    if payload.get("compare_max_chars") is None:
        payload["compare_max_chars"] = plan.get("compare_max_chars", 800)
    if payload.get("compare_titles") is None:
        payload["compare_titles"] = plan.get("compare_titles")
    if not payload.get("variants"):
        payload["variants"] = plan.get("variants") or 1
    _require_generation_sources(payload, pid)
    payload = _route_generation_or_503(payload)
    variants = int(payload.get("variants") or 1)
    payload["_variant_ids"] = reserve_variant_ids(
        project_id=pid,
        count=max(1, variants),
        explicit_variant_id=payload.get("variant_id"),
        explicit_template_id=payload.get("logic_template_id") or payload.get("logic_template"),
    )
    job_id = create_job(payload, user_id=user["id"])
    _audit("generate_async", user_id=user["id"], detail={"job_id": job_id, "project_id": pid})

    submit_isolated_job(job_id, run_legacy_generation_job, job_id, payload)
    return {"ok": True, "job_id": job_id}


@router.post("/generate_async_batch")
async def generate_async_batch(requests: list[GenerateRequest], background_tasks: BackgroundTasks, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    jobs = []
    for req in requests[:10]:
        _charge(user, COST_PER_JOB, "autoplan_generate_async_batch")
        payload = req.model_dump()
        pid = str(payload.get("project_id") or "").strip() or None
        plan = load_plan(project_id=pid) or {}
        tender = load_tender_matrix(project_id=pid) or {}
        if not payload.get("outline"):
            payload["outline"] = plan.get("outline") or []
        if not payload.get("outline"):
            payload["outline"] = tender.get("outline") or []
        if payload.get("chapter_requirements") is None:
            payload["chapter_requirements"] = plan.get("chapter_requirements") or {}
        if not payload.get("chapter_requirements"):
            payload["chapter_requirements"] = tender.get("chapter_requirements") or {}
        if payload.get("style") is None:
            payload["style"] = plan.get("style") or {}
        if not payload.get("style"):
            payload["style"] = tender.get("style") or {}
        if payload.get("chapter_pages") is None:
            payload["chapter_pages"] = plan.get("chapter_pages") or {}
        if not payload.get("chapter_pages"):
            payload["chapter_pages"] = tender.get("chapter_pages") or {}
        if payload.get("quality_strict") is None:
            payload["quality_strict"] = plan.get("quality_strict", True)
        if payload.get("auto_remediate") is None:
            payload["auto_remediate"] = plan.get("auto_remediate", True)
        if payload.get("remediate_mode") is None:
            payload["remediate_mode"] = plan.get("remediate_mode", "template")
        if payload.get("compare_mode") is None:
            payload["compare_mode"] = plan.get("compare_mode", "full")
        if payload.get("compare_max_chars") is None:
            payload["compare_max_chars"] = plan.get("compare_max_chars", 800)
        if payload.get("compare_titles") is None:
            payload["compare_titles"] = plan.get("compare_titles")
        if not payload.get("variants"):
            payload["variants"] = plan.get("variants") or 1
        _require_generation_sources(payload, pid)
        payload = _route_generation_or_503(payload)
        variants = int(payload.get("variants") or 1)
        payload["_variant_ids"] = reserve_variant_ids(
            project_id=pid,
            count=max(1, variants),
            explicit_variant_id=payload.get("variant_id"),
            explicit_template_id=payload.get("logic_template_id") or payload.get("logic_template"),
        )
        job_id = create_job(payload, user_id=user["id"])

        submit_isolated_job(job_id, run_legacy_generation_job, job_id, payload)
        jobs.append(job_id)
    return {"ok": True, "job_ids": jobs}


@router.post("/optimize")
async def optimize_content(req: OptimizeRequest, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "autoplan_optimize")
    json_path = Path("build") / "autoplan_generated.json"
    if not json_path.exists():
        raise HTTPException(status_code=404, detail="autoplan_generated.json not found")
    data = json.loads(json_path.read_text(encoding="utf-8"))
    # Keep the compatibility request fields for one cycle, but never trust its
    # provider, endpoint or credential values.  Optimization uses a fresh
    # server-owned admission receipt and the exact admitted in-memory key.
    routed = _route_generation_or_503(req.model_dump())
    _coordinator, admitted_chain = await _admit_server_provider_chain(routed)
    optimization_request = {
        "titles": list(req.titles or []),
        "instruction": str(req.instruction or ""),
        "_admitted_provider_chain": admitted_chain,
    }
    result = await optimize_sections(data, optimization_request)
    json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    docx_files = []
    if isinstance(result, dict) and isinstance(result.get("variants"), list) and result["variants"]:
        for i, variant in enumerate(result["variants"]):
            out_docx = Path("build") / f"autoplan_generated_v{i + 1}.docx"
            export_autoplan_docx(variant, str(out_docx))
            docx_files.append(str(out_docx))
    else:
        out_docx = Path("build") / "autoplan_generated_v1.docx"
        export_autoplan_docx_from_file(str(json_path), str(out_docx))
        docx_files.append(str(out_docx))
    _audit("optimize", user_id=user["id"], detail={"docx": docx_files})
    return {"ok": True, "docx": docx_files}


@router.get("/job_status")
async def job_status(job_id: str, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    job = get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    if job.get("user_id") not in (None, user["id"]):
        raise HTTPException(status_code=403, detail="forbidden")
    return {"ok": True, "job": job}


@router.get("/job_download")
async def job_download(job_id: str, kind: str = "docx", v: int = 1, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    job = get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    if job.get("user_id") not in (None, user["id"]):
        raise HTTPException(status_code=403, detail="forbidden")
    result = job.get("result") or {}
    path = result.get(kind)
    if kind in ("docx", "compare_docx") and isinstance(path, list):
        v = max(1, int(v or 1))
        path = path[v - 1] if v <= len(path) else None
    if not path:
        raise HTTPException(status_code=404, detail="file not ready")
    _audit("job_download", user_id=user["id"], detail={"job_id": job_id, "kind": kind, "v": v, "path": path})
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        if kind in ("docx", "compare_docx")
        else "application/json",
        filename=f"autoplan_{job_id}_v{v}.docx"
        if kind in ("docx", "compare_docx")
        else f"autoplan_{job_id}.{kind}",
    )


@router.get("/job_docx_versions")
async def job_docx_versions(job_id: str, kind: str = "docx", authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    job = get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    if job.get("user_id") not in (None, user["id"]):
        raise HTTPException(status_code=403, detail="forbidden")
    result = job.get("result") or {}
    docx = result.get(kind)
    if isinstance(docx, list):
        versions = list(range(1, len(docx) + 1))
        files = [Path(p).name for p in docx]
    elif isinstance(docx, str):
        versions = [1]
        files = [Path(docx).name]
    else:
        versions = []
        files = []
    return {"ok": True, "versions": versions, "files": files}


@router.get("/job_list")
async def job_list(
    limit: int = 50,
    status: str | None = None,
    has_compare: bool | None = None,
    updated_after: float | None = None,
    updated_before: float | None = None,
    recent_hours: int | None = None,
    created_after: float | None = None,
    created_before: float | None = None,
    recent_days: int | None = None,
    completed_after: float | None = None,
    completed_before: float | None = None,
    completed_recent_hours: int | None = None,
    full: bool = False,
    fields: list[str] | None = None,
    authorization: str | None = Header(default=None),
):
    user = _auth_user(authorization)
    items = list_jobs(limit=limit, user_id=user["id"])
    if status:
        items = [it for it in items if it.get("status") == status]
    if has_compare is not None:
        if has_compare:
            items = [it for it in items if it.get("result", {}).get("compare_docx")]
        else:
            items = [it for it in items if not it.get("result", {}).get("compare_docx")]
    if updated_after is not None:
        items = [it for it in items if float(it.get("updated_at") or 0) >= float(updated_after)]
    if updated_before is not None:
        items = [it for it in items if float(it.get("updated_at") or 0) <= float(updated_before)]
    if recent_hours is not None:
        import time
        threshold = time.time() - int(recent_hours) * 3600
        items = [it for it in items if float(it.get("updated_at") or 0) >= threshold]
    if created_after is not None:
        items = [it for it in items if float(it.get("created_at") or 0) >= float(created_after)]
    if created_before is not None:
        items = [it for it in items if float(it.get("created_at") or 0) <= float(created_before)]
    if recent_days is not None:
        import time
        threshold = time.time() - int(recent_days) * 86400
        items = [it for it in items if float(it.get("created_at") or 0) >= threshold]
    # completed_* 基于 updated_at（约定 status=done 时 updated_at 为完成时间）
    if completed_after is not None:
        items = [it for it in items if float(it.get("updated_at") or 0) >= float(completed_after)]
    if completed_before is not None:
        items = [it for it in items if float(it.get("updated_at") or 0) <= float(completed_before)]
    if completed_recent_hours is not None:
        import time
        threshold = time.time() - int(completed_recent_hours) * 3600
        items = [it for it in items if float(it.get("updated_at") or 0) >= threshold]
    if not full:
        _fields = _parse_fields(fields) or set(_job_list_default_fields())
        items = [{k: v for k, v in it.items() if k in _fields} for it in items]
    _audit("job_list", user_id=user["id"], detail={"count": len(items)})
    return {"ok": True, "items": items}


@router.post("/job_status_batch")
async def job_status_batch(job_ids: list[str], authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    items = []
    for jid in job_ids[:50]:
        job = get_job(jid)
        if not job:
            continue
        if job.get("user_id") not in (None, user["id"]):
            continue
        items.append(job)
    return {"ok": True, "items": items}


@router.post("/job_list_filtered")
async def job_list_filtered(req: JobListFilterRequest, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    items = list_jobs(limit=req.limit, user_id=user["id"])
    if req.job_ids:
        ids = set(req.job_ids)
        items = [it for it in items if it.get("job_id") in ids]
    if req.status:
        items = [it for it in items if it.get("status") == req.status]
    if req.has_compare is not None:
        if req.has_compare:
            items = [it for it in items if it.get("result", {}).get("compare_docx")]
        else:
            items = [it for it in items if not it.get("result", {}).get("compare_docx")]
    if req.updated_after is not None:
        items = [it for it in items if float(it.get("updated_at") or 0) >= float(req.updated_after)]
    if req.updated_before is not None:
        items = [it for it in items if float(it.get("updated_at") or 0) <= float(req.updated_before)]
    if req.recent_hours is not None:
        import time
        threshold = time.time() - int(req.recent_hours) * 3600
        items = [it for it in items if float(it.get("updated_at") or 0) >= threshold]
    if req.created_after is not None:
        items = [it for it in items if float(it.get("created_at") or 0) >= float(req.created_after)]
    if req.created_before is not None:
        items = [it for it in items if float(it.get("created_at") or 0) <= float(req.created_before)]
    if req.recent_days is not None:
        import time
        threshold = time.time() - int(req.recent_days) * 86400
        items = [it for it in items if float(it.get("created_at") or 0) >= threshold]
    if req.completed_after is not None:
        items = [it for it in items if float(it.get("updated_at") or 0) >= float(req.completed_after)]
    if req.completed_before is not None:
        items = [it for it in items if float(it.get("updated_at") or 0) <= float(req.completed_before)]
    if req.completed_recent_hours is not None:
        import time
        threshold = time.time() - int(req.completed_recent_hours) * 3600
        items = [it for it in items if float(it.get("updated_at") or 0) >= threshold]
    if not req.full:
        fields = _parse_fields(req.fields) or set(_job_list_default_fields())
        trimmed = []
        for it in items:
            trimmed.append({k: v for k, v in it.items() if k in fields})
        items = trimmed
    _audit("job_list_filtered", user_id=user["id"], detail={"count": len(items)})
    return {"ok": True, "items": items}


@router.post("/job_download_batch")
async def job_download_batch(job_ids: list[str], authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    from pathlib import Path
    import zipfile
    zip_path = Path("build") / "autoplan_batch.zip"
    if zip_path.exists():
        try:
            zip_path.unlink()
        except Exception:
            pass
    with zipfile.ZipFile(zip_path, "w") as z:
        for jid in job_ids[:20]:
            job = get_job(jid)
            if not job:
                continue
            if job.get("user_id") not in (None, user["id"]):
                continue
            result = job.get("result") or {}
            for k in ("docx", "json", "compare_docx"):
                p = result.get(k)
                if isinstance(p, list):
                    for pi in p:
                        if pi and Path(pi).exists():
                            z.write(pi, arcname=Path(pi).name)
                elif p and Path(p).exists():
                    z.write(p, arcname=Path(p).name)
    return {"ok": True, "zip": str(zip_path)}


@router.post("/job_download_compare_batch")
async def job_download_compare_batch(job_ids: list[str], authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    from pathlib import Path
    import zipfile
    zip_path = Path("build") / "autoplan_compare_batch.zip"
    if zip_path.exists():
        try:
            zip_path.unlink()
        except Exception:
            pass
    with zipfile.ZipFile(zip_path, "w") as z:
        for jid in job_ids[:20]:
            job = get_job(jid)
            if not job:
                continue
            if job.get("user_id") not in (None, user["id"]):
                continue
            result = job.get("result") or {}
            p = result.get("compare_docx")
            if isinstance(p, list):
                for pi in p:
                    if pi and Path(pi).exists():
                        z.write(pi, arcname=Path(pi).name)
            elif p and Path(p).exists():
                z.write(p, arcname=Path(p).name)
    return {"ok": True, "zip": str(zip_path)}


@router.get("/job_download_compare_batch_file")
async def job_download_compare_batch_file(authorization: str | None = Header(default=None)):
    _ = _auth_user(authorization)
    from pathlib import Path
    zip_path = Path("build") / "autoplan_compare_batch.zip"
    if not zip_path.exists():
        raise HTTPException(status_code=404, detail="zip not found")
    return FileResponse(
        str(zip_path),
        media_type="application/zip",
        filename="autoplan_compare_batch.zip",
    )


@router.get("/job_download_batch_file")
async def job_download_batch_file(authorization: str | None = Header(default=None)):
    _ = _auth_user(authorization)
    from pathlib import Path
    zip_path = Path("build") / "autoplan_batch.zip"
    if not zip_path.exists():
        raise HTTPException(status_code=404, detail="zip not found")
    return FileResponse(
        str(zip_path),
        media_type="application/zip",
        filename="autoplan_batch.zip",
    )


@router.post("/job_cleanup")
async def job_cleanup(older_than_days: int = 7, authorization: str | None = Header(default=None)):
    _ = _auth_user(authorization)
    removed = cleanup_jobs(older_than_seconds=older_than_days * 24 * 3600)
    return {"ok": True, "removed": removed}


@router.post("/export_docx")
async def export_docx(authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "autoplan_export_docx")
    json_path = Path("build") / "autoplan_generated.json"
    if not json_path.exists():
        raise HTTPException(status_code=404, detail="autoplan_generated.json not found")
    data = json.loads(json_path.read_text(encoding="utf-8"))
    docx_files = []
    if isinstance(data, dict) and isinstance(data.get("variants"), list) and data["variants"]:
        for i, variant in enumerate(data["variants"]):
            out_docx = Path("build") / f"autoplan_generated_v{i + 1}.docx"
            export_autoplan_docx(variant, str(out_docx))
            docx_files.append(str(out_docx))
    else:
        out_docx = Path("build") / "autoplan_generated_v1.docx"
        export_autoplan_docx_from_file(str(json_path), str(out_docx))
        docx_files.append(str(out_docx))
    _audit("export_docx", user_id=user["id"], detail={"docx": docx_files})
    return {"ok": True, "docx": docx_files}


@router.post("/export_compare_docx")
async def export_compare_docx(v: int = 1, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "autoplan_export_compare_docx")
    json_path = Path("build") / "autoplan_generated.json"
    if not json_path.exists():
        raise HTTPException(status_code=404, detail="autoplan_generated.json not found")
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and isinstance(data.get("variants"), list) and data["variants"]:
        v = max(1, int(v or 1))
        data = data["variants"][v - 1] if v <= len(data["variants"]) else data["variants"][0]
    out_docx = Path("build") / f"autoplan_compare_v{v}.docx"
    export_autoplan_compare_docx(data, str(out_docx))
    _audit("export_compare_docx", user_id=user["id"], detail={"v": v, "docx": str(out_docx)})
    return {"ok": True, "docx": str(out_docx)}


@router.post("/export_compare_docx_all")
async def export_compare_docx_all(authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "autoplan_export_compare_docx_all")
    json_path = Path("build") / "autoplan_generated.json"
    if not json_path.exists():
        raise HTTPException(status_code=404, detail="autoplan_generated.json not found")
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and isinstance(data.get("variants"), list) and data["variants"]:
        docx_files = []
        for i, variant in enumerate(data["variants"]):
            out_docx = Path("build") / f"autoplan_compare_v{i + 1}.docx"
            export_autoplan_compare_docx(variant, str(out_docx))
            docx_files.append(str(out_docx))
        _audit("export_compare_docx_all", user_id=user["id"], detail={"docx": docx_files})
        return {"ok": True, "docx": docx_files}
    out_docx = Path("build") / "autoplan_compare_v1.docx"
    export_autoplan_compare_docx(data, str(out_docx))
    _audit("export_compare_docx_all", user_id=user["id"], detail={"docx": [str(out_docx)]})
    return {"ok": True, "docx": [str(out_docx)]}


@router.get("/download_docx")
async def download_docx(v: int = 1, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "autoplan_download_docx")
    v = max(1, int(v or 1))
    out_docx = Path("build") / f"autoplan_generated_v{v}.docx"
    if not out_docx.exists():
        raise HTTPException(status_code=404, detail="docx not found")
    _audit("download_docx", user_id=user["id"], detail={"v": v, "path": str(out_docx)})
    return FileResponse(
        str(out_docx),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=f"autoplan_generated_v{v}.docx",
    )


@router.get("/download_compare_docx")
async def download_compare_docx(v: int = 1, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "autoplan_download_compare_docx")
    v = max(1, int(v or 1))
    out_docx = Path("build") / f"autoplan_compare_v{v}.docx"
    if not out_docx.exists():
        raise HTTPException(status_code=404, detail="docx not found")
    _audit("download_compare_docx", user_id=user["id"], detail={"v": v, "path": str(out_docx)})
    return FileResponse(
        str(out_docx),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename=f"autoplan_compare_v{v}.docx",
    )


@router.get("/download_compare_docx_all")
async def download_compare_docx_all(authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "autoplan_download_compare_docx_all")
    from zipfile import ZipFile
    build_dir = Path("build")
    files = sorted(build_dir.glob("autoplan_compare_v*.docx"))
    if not files:
        raise HTTPException(status_code=404, detail="docx not found")
    zip_path = build_dir / "autoplan_compare_all.zip"
    if zip_path.exists():
        try:
            zip_path.unlink()
        except Exception:
            pass
    with ZipFile(str(zip_path), "w") as z:
        for p in files:
            z.write(p, arcname=p.name)
    _audit("download_compare_docx_all", user_id=user["id"], detail={"zip": str(zip_path)})
    return FileResponse(
        str(zip_path),
        media_type="application/zip",
        filename="autoplan_compare_all.zip",
    )


@router.get("/compare_docx_versions")
async def compare_docx_versions(authorization: str | None = Header(default=None)):
    _ = _auth_user(authorization)
    from pathlib import Path
    files = sorted(Path("build").glob("autoplan_compare_v*.docx"))
    versions = []
    for p in files:
        name = p.stem  # autoplan_compare_vX
        if "_v" in name:
            v = name.split("_v")[-1]
            if v.isdigit():
                versions.append(int(v))
    versions = sorted(set(versions))
    return {
        "ok": True,
        "versions": versions,
        "files": [f"autoplan_compare_v{v}.docx" for v in versions],
    }


@router.get("/job_list_fields")
async def job_list_fields(authorization: str | None = Header(default=None)):
    _ = _auth_user(authorization)
    return {
        "ok": True,
        "default_fields": sorted(_job_list_default_fields()),
        "field_alias": {k: sorted(list(v)) for k, v in _job_list_field_alias().items()},
    }


@router.get("/audit")
async def audit_log(
    limit: int = 200,
    action: str | None = None,
    user_id: int | None = None,
    ts_after: str | None = None,
    ts_before: str | None = None,
    authorization: str | None = Header(default=None),
):
    user = _auth_user(authorization)
    uid = user_id
    if user_id is not None and not _is_admin(authorization):
        raise HTTPException(status_code=403, detail="forbidden")
    if uid is None and not _is_admin(authorization):
        uid = user["id"]
    items = _read_audit(limit=limit, action=action, user_id=uid, ts_after=ts_after, ts_before=ts_before)
    return {"ok": True, "items": items}


@router.get("/audit/summary")
async def audit_summary(
    limit: int = 10000,
    by_user: bool = False,
    authorization: str | None = Header(default=None),
):
    user = _auth_user(authorization)
    uid = None if _is_admin(authorization) else user["id"]
    summary = _audit_summary(limit=limit, user_id=uid, by_user=by_user and _is_admin(authorization))
    return {"ok": True, "summary": summary}


@router.get("/audit/stats")
async def audit_stats(
    days: int = 30,
    authorization: str | None = Header(default=None),
):
    user = _auth_user(authorization)
    uid = None if _is_admin(authorization) else user["id"]
    stats = _audit_stats_by_day(limit_days=days, user_id=uid)
    return {"ok": True, "stats": stats}


@router.post("/audit/query")
async def audit_query(req: AuditQueryRequest, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    uid = req.user_id
    if req.user_id is not None and not _is_admin(authorization):
        raise HTTPException(status_code=403, detail="forbidden")
    if uid is None and not _is_admin(authorization):
        uid = user["id"]
    items = _read_audit(
        limit=req.limit,
        action=req.action,
        user_id=uid,
        ts_after=req.ts_after,
        ts_before=req.ts_before,
    )
    return {"ok": True, "items": items}


@router.get("/audit/export")
async def audit_export(
    fmt: str = "json",
    limit: int = 1000,
    action: str | None = None,
    user_id: int | None = None,
    ts_after: str | None = None,
    ts_before: str | None = None,
    authorization: str | None = Header(default=None),
):
    user = _auth_user(authorization)
    uid = user_id
    if user_id is not None and not _is_admin(authorization):
        raise HTTPException(status_code=403, detail="forbidden")
    if uid is None and not _is_admin(authorization):
        uid = user["id"]
    items = _read_audit(limit=limit, action=action, user_id=uid, ts_after=ts_after, ts_before=ts_before)
    fmt = (fmt or "json").lower()
    if fmt not in ("json", "csv"):
        raise HTTPException(status_code=400, detail="format must be json|csv")
    if fmt == "json":
        from fastapi.responses import JSONResponse
        return JSONResponse(content={"ok": True, "items": items})
    # csv
    import csv
    import io
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ts", "action", "user_id", "detail"])
    for it in items:
        writer.writerow([it.get("ts"), it.get("action"), it.get("user_id"), json.dumps(it.get("detail"), ensure_ascii=False)])
    from fastapi.responses import PlainTextResponse
    return PlainTextResponse(
        output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=audit_export.csv"},
    )


@router.get("/audit/export_xlsx")
async def audit_export_xlsx(
    limit: int = 1000,
    action: str | None = None,
    user_id: int | None = None,
    ts_after: str | None = None,
    ts_before: str | None = None,
    authorization: str | None = Header(default=None),
):
    user = _auth_user(authorization)
    uid = user_id
    if user_id is not None and not _is_admin(authorization):
        raise HTTPException(status_code=403, detail="forbidden")
    if uid is None and not _is_admin(authorization):
        uid = user["id"]
    items = _read_audit(limit=limit, action=action, user_id=uid, ts_after=ts_after, ts_before=ts_before)
    import io
    output = io.BytesIO()
    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "audit"
    headers = ["ts", "action", "user_id", "detail"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, it in enumerate(items, start=2):
        ws.cell(row=r, column=1, value=it.get("ts"))
        ws.cell(row=r, column=2, value=it.get("action"))
        ws.cell(row=r, column=3, value=it.get("user_id"))
        ws.cell(row=r, column=4, value=json.dumps(it.get("detail"), ensure_ascii=False))
    wb.save(output)
    output.seek(0)
    from fastapi.responses import Response
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=audit_export.xlsx"},
    )


@router.get("/audit/export_file_download")
async def audit_export_file_download(
    filename: str = "audit_export",
    user_id: int | None = None,
    authorization: str | None = Header(default=None),
):
    user = _auth_user(authorization)
    uid = user_id
    if user_id is not None and not _is_admin(authorization):
        raise HTTPException(status_code=403, detail="forbidden")
    if uid is None and not _is_admin(authorization):
        uid = user["id"]
    base_dir = Path("build") / "audit_exports" / str(uid or user["id"])
    json_path = base_dir / f"{filename}.json"
    csv_path = base_dir / f"{filename}.csv"
    xlsx_path = base_dir / f"{filename}.xlsx"
    if json_path.exists():
        return FileResponse(str(json_path), media_type="application/json", filename=f"{filename}.json")
    if csv_path.exists():
        return FileResponse(str(csv_path), media_type="text/csv", filename=f"{filename}.csv")
    if xlsx_path.exists():
        return FileResponse(
            str(xlsx_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{filename}.xlsx",
        )
    raise HTTPException(status_code=404, detail="export file not found")


@router.get("/audit/export_file_list")
async def audit_export_file_list(
    user_id: int | None = None,
    limit: int = 100,
    authorization: str | None = Header(default=None),
):
    user = _auth_user(authorization)
    uid = user_id
    if user_id is not None and not _is_admin(authorization):
        raise HTTPException(status_code=403, detail="forbidden")
    if uid is None and not _is_admin(authorization):
        uid = user["id"]
    base_dir = Path("build") / "audit_exports" / str(uid or user["id"])
    if not base_dir.exists():
        return {"ok": True, "files": []}
    files = sorted(base_dir.glob("*.*"), key=lambda p: p.stat().st_mtime, reverse=True)
    items = []
    for p in files[:limit]:
        items.append(
            {
                "name": p.name,
                "path": str(p),
                "size": p.stat().st_size,
                "mtime": p.stat().st_mtime,
            }
        )
    return {"ok": True, "files": items}


@router.post("/audit/export_file_list")
async def audit_export_file_list_post(req: AuditExportListRequest, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    uid = req.user_id
    if req.user_id is not None and not _is_admin(authorization):
        raise HTTPException(status_code=403, detail="forbidden")
    if uid is None and not _is_admin(authorization):
        uid = user["id"]
    base_dir = Path("build") / "audit_exports" / str(uid or user["id"])
    if not base_dir.exists():
        return {"ok": True, "files": []}
    files = sorted(base_dir.glob("*.*"), key=lambda p: p.stat().st_mtime, reverse=True)
    items = []
    for p in files[: req.limit]:
        items.append(
            {
                "name": p.name,
                "path": str(p),
                "size": p.stat().st_size,
                "mtime": p.stat().st_mtime,
            }
        )
    return {"ok": True, "files": items}


@router.post("/audit/export_file_cleanup")
async def audit_export_file_cleanup(req: AuditExportCleanupRequest, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    uid = req.user_id
    if req.user_id is not None and not _is_admin(authorization):
        raise HTTPException(status_code=403, detail="forbidden")
    if uid is None and not _is_admin(authorization):
        uid = user["id"]
    if req.older_than_days is None and req.keep_latest_n is None:
        raise HTTPException(status_code=400, detail="set older_than_days or keep_latest_n")
    import time
    base_dir = Path("build") / "audit_exports" / str(uid or user["id"])
    if not base_dir.exists():
        return {"ok": True, "removed": 0}
    files = sorted(base_dir.glob("*.*"), key=lambda p: p.stat().st_mtime, reverse=True)
    removed = 0
    now = time.time()
    if req.older_than_days is not None:
        threshold = now - req.older_than_days * 86400
        for p in files:
            if p.stat().st_mtime < threshold:
                try:
                    p.unlink()
                    removed += 1
                except Exception:
                    pass
    if req.keep_latest_n is not None and req.keep_latest_n >= 0:
        for p in files[req.keep_latest_n:]:
            try:
                if p.exists():
                    p.unlink()
                    removed += 1
            except Exception:
                pass
    return {"ok": True, "removed": removed}


@router.post("/audit/export_file")
async def audit_export_file(req: AuditExportRequest, authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    uid = req.user_id
    if req.user_id is not None and not _is_admin(authorization):
        raise HTTPException(status_code=403, detail="forbidden")
    if uid is None and not _is_admin(authorization):
        uid = user["id"]
    items = _read_audit(
        limit=req.limit,
        action=req.action,
        user_id=uid,
        ts_after=req.ts_after,
        ts_before=req.ts_before,
    )
    fmt = (req.fmt or "json").lower()
    if fmt not in ("json", "csv", "xlsx"):
        raise HTTPException(status_code=400, detail="format must be json|csv|xlsx")
    build_dir = Path("build") / "audit_exports" / str(uid or user["id"])
    build_dir.mkdir(parents=True, exist_ok=True)
    base_name = req.filename or "audit_export"
    if fmt == "json":
        out_path = build_dir / f"{base_name}.json"
        out_path.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
        return {"ok": True, "file": str(out_path)}
    if fmt == "csv":
        import csv
        out_path = build_dir / f"{base_name}.csv"
        with out_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["ts", "action", "user_id", "detail"])
            for it in items:
                writer.writerow([it.get("ts"), it.get("action"), it.get("user_id"), json.dumps(it.get("detail"), ensure_ascii=False)])
        return {"ok": True, "file": str(out_path)}
    # xlsx
    out_path = build_dir / f"{base_name}.xlsx"
    try:
        from openpyxl import Workbook
    except Exception:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "audit"
    headers = ["ts", "action", "user_id", "detail"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, it in enumerate(items, start=2):
        ws.cell(row=r, column=1, value=it.get("ts"))
        ws.cell(row=r, column=2, value=it.get("action"))
        ws.cell(row=r, column=3, value=it.get("user_id"))
        ws.cell(row=r, column=4, value=json.dumps(it.get("detail"), ensure_ascii=False))
    wb.save(str(out_path))
    return {"ok": True, "file": str(out_path)}


@router.get("/download_docx_all")
async def download_docx_all(authorization: str | None = Header(default=None)):
    user = _auth_user(authorization)
    _charge(user, COST_PER_JOB, "autoplan_download_docx_all")
    from zipfile import ZipFile
    build_dir = Path("build")
    files = sorted(build_dir.glob("autoplan_generated_v*.docx"))
    if not files:
        raise HTTPException(status_code=404, detail="docx not found")
    zip_path = build_dir / "autoplan_generated_all.zip"
    if zip_path.exists():
        try:
            zip_path.unlink()
        except Exception:
            pass
    with ZipFile(str(zip_path), "w") as z:
        for p in files:
            z.write(p, arcname=p.name)
    return FileResponse(
        str(zip_path),
        media_type="application/zip",
        filename="autoplan_generated_all.zip",
    )


@router.get("/docx_versions")
async def docx_versions(authorization: str | None = Header(default=None)):
    _ = _auth_user(authorization)
    from pathlib import Path
    files = sorted(Path("build").glob("autoplan_generated_v*.docx"))
    versions = []
    for p in files:
        name = p.stem  # autoplan_generated_vX
        if "_v" in name:
            v = name.split("_v")[-1]
            if v.isdigit():
                versions.append(int(v))
    versions = sorted(set(versions))
    return {
        "ok": True,
        "versions": versions,
        "files": [f"autoplan_generated_v{v}.docx" for v in versions],
    }


@router.get("/generate_status")
async def generate_status(authorization: str | None = Header(default=None)):
    _ = _auth_user(authorization)
    json_path = Path("build") / "autoplan_generated.json"
    v1_path = Path("build") / "autoplan_generated_v1.docx"
    any_docx = sorted(Path("build").glob("autoplan_generated_v*.docx"))
    status = {
        "json_exists": json_path.exists(),
        "docx_exists": bool(any_docx),
        "json_mtime": json_path.stat().st_mtime if json_path.exists() else None,
        "docx_mtime": v1_path.stat().st_mtime if v1_path.exists() else (any_docx[0].stat().st_mtime if any_docx else None),
        "docx_versions": [p.name for p in any_docx],
    }
    return {"ok": True, "status": status}


@router.get("/check")
async def check_generated(authorization: str | None = Header(default=None)):
    _ = _auth_user(authorization)
    json_path = Path("build") / "autoplan_generated.json"
    if not json_path.exists():
        raise HTTPException(status_code=404, detail="autoplan_generated.json not found")
    data = json.loads(json_path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and isinstance(data.get("variants"), list) and data["variants"]:
        return {"ok": True, "quality": [v.get("quality_checks") for v in data["variants"]]}
    return {"ok": True, "quality": data.get("quality_checks")}
