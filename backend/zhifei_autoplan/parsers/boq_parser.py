from __future__ import annotations

import asyncio
import json
import math
import re
from pathlib import Path
from typing import Any, ClassVar

import pdfplumber

from backend.zhifei_autoplan.models import BoQItem, ConstructionProcess, Resource


class BoQParser:
    """
    清单智能识别与统计引擎（Module 2）
    - MECE 原则：清单项按“材料/工序/资源”三段映射，互不重复但覆盖所有项目
    """

    _HEADER_ALIASES: ClassVar[dict[str, tuple[str, ...]]] = {
        "code": (
            "序号",
            "编号",
            "清单编码",
            "清单项目编码",
            "项目编码",
            "清单号",
            "子目编码",
            "编码",
        ),
        "name": (
            "项目名称",
            "清单项目名称",
            "分部分项工程名称",
            "子目名称",
            "名称",
            "项目",
        ),
        "project_feature": (
            "项目特征描述",
            "项目特征",
            "特征描述",
            "工作内容及特征",
        ),
        "unit": ("计量单位", "单位名称", "单位"),
        "qty": ("工程量", "清单工程量", "计量数量", "合计数量", "数量"),
        "unit_price": ("综合单价", "材料单价", "综合价", "单价", "unit_price"),
        "total_price": ("合价", "总价", "金额", "合计", "total_price"),
    }

    @staticmethod
    def _clean_text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()

    @classmethod
    def _normalize_header(cls, value: Any) -> str:
        text = cls._clean_text(value).lower()
        return re.sub(r"[\s:：()（）\[\]【】]+", "", text)

    @classmethod
    def _header_map(cls, cells: list[Any] | tuple[Any, ...]) -> dict[str, int]:
        aliases = {
            key: {cls._normalize_header(alias) for alias in values}
            for key, values in cls._HEADER_ALIASES.items()
        }
        result: dict[str, int] = {}
        for index, cell in enumerate(cells):
            normalized = cls._normalize_header(cell)
            if not normalized:
                continue
            for field, names in aliases.items():
                if normalized in names and field not in result:
                    result[field] = index
                    break
        return result

    @staticmethod
    def _is_usable_header(mapping: dict[str, int]) -> bool:
        core = all(field in mapping for field in ("name", "unit", "qty"))
        signature = any(
            field in mapping
            for field in ("code", "project_feature", "unit_price", "total_price")
        )
        return core and signature

    @staticmethod
    def _safe_locator_name(path: str) -> str:
        name = Path(str(path or "boq")).name or "boq"
        return re.sub(r"[^0-9A-Za-z._\-\u4e00-\u9fff]+", "_", name).strip("_") or "boq"

    @classmethod
    def _source_locator(
        cls,
        path: str,
        *,
        page: int | None = None,
        sheet_name: str | None = None,
        table_index: int | None = None,
        row_index: int | None = None,
    ) -> dict[str, Any]:
        file_name = Path(str(path or "boq")).name or "boq"
        parts = [f"boq:{cls._safe_locator_name(path)}"]
        result: dict[str, Any] = {"file_name": file_name}
        if page is not None:
            result["page"] = int(page)
            parts.append(f"p{int(page)}")
        if sheet_name:
            result["sheet_name"] = str(sheet_name)
            parts.append(f"sheet-{re.sub(r'[^0-9A-Za-z_\-\u4e00-\u9fff]+', '_', str(sheet_name))}")
        if table_index is not None:
            result["table_index"] = int(table_index)
            parts.append(f"t{int(table_index)}")
        if row_index is not None:
            result["row_index"] = int(row_index)
            parts.append(f"r{int(row_index)}")
        result["locator"] = "/".join(parts)
        return result

    @staticmethod
    def _cell(row: list[Any] | tuple[Any, ...], index: int | None) -> Any:
        if index is None or index < 0 or index >= len(row):
            return None
        return row[index]

    @classmethod
    def _mapped_row(
        cls,
        row: list[Any] | tuple[Any, ...],
        mapping: dict[str, int],
        *,
        source_locator: dict[str, Any],
    ) -> dict[str, Any]:
        return {
            "code": cls._clean_text(cls._cell(row, mapping.get("code"))),
            "name": cls._clean_text(cls._cell(row, mapping.get("name"))),
            "project_feature": cls._clean_text(
                cls._cell(row, mapping.get("project_feature"))
            ),
            "qty": cls._cell(row, mapping.get("qty")),
            "unit": cls._clean_text(cls._cell(row, mapping.get("unit"))),
            # Prices are deliberately absent unless an explicit price header
            # exists.  In particular, a five-column unpriced BoQ must never
            # reinterpret its engineering quantity as a unit price.
            "unit_price": cls._cell(row, mapping.get("unit_price")),
            "total_price": cls._cell(row, mapping.get("total_price")),
            "source_locator": source_locator,
        }

    async def parse(self, path: str) -> tuple[list[BoQItem], dict[str, Any]]:
        # PDF/Excel extraction and the subsequent row-to-model normalization are
        # both CPU-bound.  Keep the complete pipeline off FastAPI's event-loop
        # thread so lightweight endpoints (notably /health and job polling)
        # remain responsive while a large BoQ is being parsed.
        return await asyncio.to_thread(self._parse_sync, path)

    def _parse_sync(self, path: str) -> tuple[list[BoQItem], dict[str, Any]]:
        if path.lower().endswith((".xlsx", ".xls")):
            rows = self._read_excel(path)
        elif path.lower().endswith(".pdf"):
            rows = self._read_pdf_tables(path)
        else:
            rows = []

        items = [self._row_to_item(row) for row in rows if self._is_leaf_row(row)]
        stats = self._calc_stats(items)
        return items, stats

    def _read_excel(self, path: str) -> list[dict[str, Any]]:
        """
        Read BoQ Excel without pandas to avoid heavy C-extension dependencies.
        Supports:
        - .xlsx via openpyxl
        - .xls via xlrd (optional; if not installed, returns empty)
        """
        p = str(path or "")
        if not p:
            return []
        lower = p.lower()
        if lower.endswith(".xls") and not lower.endswith(".xlsx"):
            return self._read_xls_xlrd(p)
        return self._read_xlsx_openpyxl(p)

    def _read_xlsx_openpyxl(self, path: str) -> list[dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return []
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
        # Vendor workbooks can fail in several parser layers (ZIP, XML, or
        # openpyxl validation); every such failure is a fail-closed empty read.
        except Exception:  # noqa: BLE001
            return []
        ws = wb.active

        header_row_idx = None
        header_mapping: dict[str, int] = {}
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx > 30:
                break
            if not row:
                continue
            cells = [self._clean_text(c) for c in row]
            if not any(cells):
                continue
            mapping = self._header_map(cells)
            if self._is_usable_header(mapping):
                header_row_idx = r_idx
                header_mapping = mapping
                break

        if header_row_idx is None:
            return []

        rows: list[dict[str, Any]] = []
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
            start=header_row_idx + 1,
        ):
            if not row:
                continue
            payload = self._mapped_row(
                row,
                header_mapping,
                source_locator=self._source_locator(
                    path,
                    sheet_name=str(ws.title or "Sheet"),
                    row_index=row_idx,
                ),
            )
            if not any(
                payload.get(key)
                for key in ("code", "name", "qty", "unit", "unit_price", "total_price")
            ):
                continue
            rows.append(payload)
        return rows

    def _read_xls_xlrd(self, path: str) -> list[dict[str, Any]]:
        """
        Optional .xls support (requires xlrd). If unavailable, returns empty.
        """
        try:
            import xlrd  # type: ignore
        except ImportError:
            return []
        try:
            book = xlrd.open_workbook(path)
            sheet = book.sheet_by_index(0)
        # The optional xlrd parser exposes format-specific exception classes;
        # malformed input consistently fails closed as an empty read.
        except Exception:  # noqa: BLE001
            return []

        header_row_idx = None
        header_mapping: dict[str, int] = {}
        for r in range(min(30, sheet.nrows)):
            cells = [self._clean_text(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            if not any(cells):
                continue
            mapping = self._header_map(cells)
            if self._is_usable_header(mapping):
                header_row_idx = r
                header_mapping = mapping
                break

        if header_row_idx is None:
            return []

        rows: list[dict[str, Any]] = []
        for r in range(header_row_idx + 1, sheet.nrows):
            row = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            payload = self._mapped_row(
                row,
                header_mapping,
                source_locator=self._source_locator(
                    path,
                    sheet_name=str(sheet.name or "Sheet"),
                    row_index=r + 1,
                ),
            )
            if not any(
                payload.get(key)
                for key in ("code", "name", "qty", "unit", "unit_price", "total_price")
            ):
                continue
            rows.append(payload)
        return rows

    def _read_pdf_tables(self, path: str) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        with pdfplumber.open(path) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables() or []
                for table_number, table in enumerate(tables, start=1):
                    if not table:
                        continue
                    header_index: int | None = None
                    header_mapping: dict[str, int] = {}
                    for candidate_index, candidate in enumerate(table[:6]):
                        if not candidate:
                            continue
                        mapping = self._header_map(candidate)
                        if self._is_usable_header(mapping):
                            header_index = candidate_index
                            header_mapping = mapping
                            break
                    # Never infer semantics from column count.  The historical
                    # four-column BoQ signature (编码/名称/数量/单位 and aliases)
                    # is already proven by ``_header_map`` above; an unrelated
                    # four-column table must not be reinterpreted as a BoQ.
                    if header_index is None:
                        continue
                    for row_number, row in enumerate(
                        table[header_index + 1 :],
                        start=header_index + 2,
                    ):
                        if not row:
                            continue
                        required_last_index = max(
                            header_mapping[field] for field in ("name", "unit", "qty")
                        )
                        if len(row) <= required_last_index:
                            continue
                        payload = self._mapped_row(
                            row,
                            header_mapping,
                            source_locator=self._source_locator(
                                path,
                                page=page_number,
                                table_index=table_number,
                                row_index=row_number,
                            ),
                        )
                        if payload.get("name"):
                            rows.append(payload)
        return rows

    def _row_to_item(self, r: dict[str, Any]) -> BoQItem:
        unit = self._normalize_unit(r.get("unit"))
        qty = self._to_quantity(r.get("qty"), unit=unit)
        unit_price = self._to_float(r.get("unit_price"))
        total_price = self._to_float(r.get("total_price"))
        if total_price is None and qty is not None and unit_price is not None:
            total_price = qty * unit_price
        item = BoQItem(
            boq_code=r.get("code") or "",
            name=r.get("name") or "",
            project_feature=r.get("project_feature") or None,
            quantity=qty,
            unit=unit or None,
            unit_price=unit_price,
            total_price=total_price,
            source_locator=(
                dict(r.get("source_locator"))
                if isinstance(r.get("source_locator"), dict)
                else None
            ),
        )
        process, resources = self.map_boq_to_process(item)
        item.process = process
        item.resources = resources
        return item

    @staticmethod
    def _normalize_unit(value: Any) -> str:
        return re.sub(r"\s+", "", str(value or "")).strip()

    @classmethod
    def _valid_unit(cls, value: Any) -> bool:
        unit = cls._normalize_unit(value)
        if not unit or len(unit) > 24:
            return False
        return re.fullmatch(r"[0-9A-Za-z%./·²³㎎㎡㎥一-鿿()（）\-]+", unit) is not None

    @classmethod
    def _to_quantity(cls, value: Any, *, unit: str = "") -> float | None:
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            number = float(value)
            return number if math.isfinite(number) and number >= 0 else None
        text = str(value).strip().replace(",", "").replace("，", "")
        if not text:
            return None
        match = re.fullmatch(
            r"\+?(\d+(?:\.\d+)?|\.\d+)(?:\s*([^\d\s].*))?",
            text,
        )
        if not match:
            return None
        suffix = cls._normalize_unit(match.group(2) or "")
        normalized_unit = cls._normalize_unit(unit)
        if suffix and normalized_unit and suffix.lower() != normalized_unit.lower():
            return None
        try:
            number = float(match.group(1))
        except (TypeError, ValueError):
            return None
        return number if math.isfinite(number) and number >= 0 else None

    @classmethod
    def _is_leaf_row(cls, row: dict[str, Any]) -> bool:
        if not str(row.get("name") or "").strip():
            return False
        unit = cls._normalize_unit(row.get("unit"))
        if not cls._valid_unit(unit):
            return False
        return cls._to_quantity(row.get("qty"), unit=unit) is not None

    def map_boq_to_process(self, boq_item: BoQItem) -> tuple[ConstructionProcess, list[Resource]]:
        # 三维映射逻辑：清单 -> 工序 -> 资源
        name = boq_item.name
        mapping = self._load_process_rules() or [
            (["混凝土"], "混凝土浇筑", ["搅拌车", "泵车"]),
            (["钢筋"], "钢筋绑扎", ["钢筋工", "电焊机"]),
            (["模板"], "模板安装", ["模板工", "支撑材料"]),
            (["回填"], "土方回填", ["压路机", "装载机"]),
            (["管道"], "管道安装", ["挖机", "起重机"]),
            # Extra common rules (kept after the core five to preserve first-match behavior).
            (["防水", "卷材"], "防水施工", ["防水工", "热风焊机"]),
            (["砌体", "砌筑", "加气块", "砖"], "砌体砌筑", ["砌筑工", "砂浆搅拌机"]),
            (["抹灰", "粉刷"], "抹灰工程", ["抹灰工", "砂浆搅拌机"]),
            (["保温"], "保温施工", ["保温工", "切割机"]),
            (["电缆", "桥架"], "电气安装", ["电工", "电缆牵引机"]),
            (["消防"], "消防安装", ["管道工", "电工"]),
            (["通风", "空调", "暖通"], "暖通安装", ["管道工", "焊工"]),
            (["门窗"], "门窗安装", ["安装工", "电钻"]),
            (["幕墙"], "幕墙安装", ["安装工", "吊篮"]),
            (["沥青"], "沥青路面施工", ["摊铺机", "压路机"]),
            (["路基"], "路基施工", ["推土机", "压路机"]),
        ]

        proc = ConstructionProcess(name="通用施工工序", standard=None, risks=[])
        res: list[Resource] = []
        for kws, proc_name, res_names in mapping:
            if any(str(kw) in (name or "") for kw in (kws or [])):
                proc = ConstructionProcess(name=str(proc_name), standard=None, risks=[])
                res = [Resource(name=str(r)) for r in (res_names or [])]
                break
        return proc, res

    _PROCESS_RULES_CACHE: list[tuple[list[str], str, list[str]]] | None = None
    _PROCESS_RULES_MTIME_NS: int | None = None

    def _load_process_rules(self) -> list[tuple[list[str], str, list[str]]] | None:
        """
        Optional user-configurable mapping rules.
        File: backend/data/autoplan/boq_process_rules.json
        Format:
          {
            "rules": [
              {"match": ["混凝土", "现浇"], "process": "混凝土浇筑", "resources": ["搅拌车", "泵车"]},
              ...
            ]
          }
        """
        cfg = Path("backend/data/autoplan/boq_process_rules.json")
        if not cfg.exists() or not cfg.is_file():
            return None
        try:
            mtime = int(cfg.stat().st_mtime_ns)
        except OSError:
            mtime = None
        if self._PROCESS_RULES_CACHE is not None and mtime is not None and self._PROCESS_RULES_MTIME_NS == mtime:
            return self._PROCESS_RULES_CACHE

        try:
            obj = json.loads(cfg.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return None
        if not isinstance(obj, dict):
            return None
        rules = obj.get("rules")
        if not isinstance(rules, list) or not rules:
            return None

        parsed: list[tuple[list[str], str, list[str]]] = []
        for r in rules:
            if not isinstance(r, dict):
                continue
            raw_match = r.get("match") or r.get("keywords") or r.get("kw") or []
            if isinstance(raw_match, str):
                raw_match = [raw_match]
            if not isinstance(raw_match, list):
                continue
            kws = [str(x).strip() for x in raw_match if str(x).strip()]
            proc = str(r.get("process") or r.get("process_name") or "").strip()
            raw_res = r.get("resources") or r.get("res") or []
            if isinstance(raw_res, str):
                raw_res = [raw_res]
            res = [str(x).strip() for x in (raw_res or []) if str(x).strip()] if isinstance(raw_res, list) else []
            if kws and proc:
                parsed.append((kws, proc, res))
        if not parsed:
            return None
        self._PROCESS_RULES_CACHE = parsed
        self._PROCESS_RULES_MTIME_NS = mtime
        return parsed

    def _split_code_hierarchy(self, code: str) -> tuple[str, str]:
        raw = str(code or "").strip()
        if not raw:
            return "", ""

        # Prefer explicit separators first.
        if any(sep in raw for sep in (".", "-", "_", "/")):
            parts = [p for p in re.split(r"[.\-_/]+", raw) if p]
            if len(parts) == 1:
                return parts[0], parts[0]
            return parts[0], ".".join(parts[:2])

        # Fallback for long continuous digit codes.
        compact = re.sub(r"[^0-9A-Za-z]+", "", raw)
        if compact.isdigit() and len(compact) >= 4:
            return compact[:2], compact[:4]
        if len(compact) >= 2:
            return compact[:2], compact[:4] if len(compact) >= 4 else compact[:2]
        return compact, compact

    def _infer_chapter_structure(self, items: list[BoQItem]) -> dict[str, Any]:
        by_chapter: dict[str, dict[str, Any]] = {}
        by_item: dict[str, dict[str, Any]] = {}

        for idx, it in enumerate(items, start=1):
            code = str(it.boq_code or "").strip()
            chapter_id, subchapter_id = self._split_code_hierarchy(code)
            if not chapter_id:
                # Fallback: infer chapter by process name or first two chars in name.
                pname = str((it.process.name if it.process else "") or "").strip()
                if pname:
                    chapter_id = pname
                else:
                    chapter_id = str(it.name or f"CH-{idx}")[:2] or f"CH-{idx}"
            if not subchapter_id:
                subchapter_id = chapter_id

            chapter = by_chapter.setdefault(
                chapter_id,
                {
                    "chapter_id": chapter_id,
                    "item_count": 0,
                    "total_quantity": 0.0,
                    "subchapters": {},
                },
            )
            chapter["item_count"] += 1
            chapter["total_quantity"] += float(it.quantity or 0.0)

            sub = chapter["subchapters"].setdefault(
                subchapter_id,
                {
                    "subchapter_id": subchapter_id,
                    "item_count": 0,
                    "total_quantity": 0.0,
                },
            )
            sub["item_count"] += 1
            sub["total_quantity"] += float(it.quantity or 0.0)

            item_key = code or str(it.name or f"item_{idx}")
            by_item[item_key] = {
                "chapter_id": chapter_id,
                "subchapter_id": subchapter_id,
            }

        chapters: list[dict[str, Any]] = []
        for cid in sorted(by_chapter.keys()):
            chapter = by_chapter[cid]
            children = []
            for sid in sorted(chapter["subchapters"].keys()):
                children.append(chapter["subchapters"][sid])
            chapters.append(
                {
                    "chapter_id": chapter["chapter_id"],
                    "item_count": int(chapter["item_count"]),
                    "total_quantity": round(float(chapter["total_quantity"]), 6),
                    "subchapter_count": len(children),
                    "subchapters": children,
                }
            )

        return {
            "chapter_count": len(chapters),
            "chapters": chapters,
            "by_item": by_item,
        }

    def _calc_complexity_metrics(self, items: list[BoQItem]) -> dict[str, Any]:
        if not items:
            return {
                "complexity_index": 0.0,
                "resource_density_index": 0.0,
                "quantity_scale_index": 0.0,
                "construction_density_index": 0.0,
                "process_diversity": 0.0,
            }

        process_names = {str(it.process.name) for it in items if it.process and str(it.process.name).strip()}
        process_diversity = len(process_names) / max(1, len(items))

        resource_counts = [len(it.resources or []) for it in items]
        avg_resource_count = sum(resource_counts) / max(1, len(resource_counts))
        resource_density_index = min(1.0, avg_resource_count / 4.0)

        quantities = [float(it.quantity or 0.0) for it in items if it.quantity is not None]
        total_quantity = sum(quantities)
        avg_quantity = (total_quantity / max(1, len(quantities))) if quantities else 0.0
        variance = 0.0
        if quantities:
            variance = sum((q - avg_quantity) ** 2 for q in quantities) / max(1, len(quantities))
        std_dev = math.sqrt(variance)
        dispersion = (std_dev / max(1.0, abs(avg_quantity))) if avg_quantity else 0.0
        quantity_scale_index = min(1.0, math.log10(max(1.0, total_quantity) + 1.0) / 6.0)

        # Density reflects workload concentration; higher total and lower item spread => denser.
        construction_density_index = min(1.0, (total_quantity / max(1.0, len(items))) / 1500.0)

        special_keywords = ("高性能", "特种", "抗渗", "防腐", "不锈钢", "高强", "预应力", "四新")
        hazard_keywords = ("危化", "易燃", "易爆", "有毒", "溶剂", "油漆", "涂料", "沥青", "气瓶")
        special_ratio = sum(1 for it in items if any(k in (it.name or "") for k in special_keywords)) / max(1, len(items))
        hazard_ratio = sum(1 for it in items if any(k in (it.name or "") for k in hazard_keywords)) / max(1, len(items))

        complexity_index = (
            process_diversity * 0.25
            + resource_density_index * 0.2
            + min(1.0, dispersion) * 0.2
            + construction_density_index * 0.15
            + special_ratio * 0.1
            + hazard_ratio * 0.1
        )
        complexity_index = round(min(1.0, complexity_index), 4)

        return {
            "complexity_index": complexity_index,
            "resource_density_index": round(min(1.0, resource_density_index), 4),
            "quantity_scale_index": round(min(1.0, quantity_scale_index), 4),
            "construction_density_index": round(min(1.0, construction_density_index), 4),
            "process_diversity": round(min(1.0, process_diversity), 4),
            "quantity_dispersion": round(max(0.0, dispersion), 4),
            "avg_resource_count": round(avg_resource_count, 4),
        }

    def _calc_stats(self, items: list[BoQItem]) -> dict[str, Any]:
        # 统计分析：数量级与密度（简化版）
        total_qty = 0.0
        count = 0.0
        for it in items:
            if it.quantity is not None:
                total_qty += it.quantity
                count += 1.0
        density = total_qty / count if count else 0.0
        top_quantity = sorted(
            [it for it in items if it.quantity is not None],
            key=lambda x: float(x.quantity or 0),
            reverse=True,
        )[:8]
        top_unit_price = sorted(
            [it for it in items if it.unit_price is not None],
            key=lambda x: float(x.unit_price or 0),
            reverse=True,
        )[:8]
        top_total_price = sorted(
            [it for it in items if it.total_price is not None],
            key=lambda x: float(x.total_price or 0),
            reverse=True,
        )[:8]
        material_keywords = (
            "钢筋",
            "混凝土",
            "水泥",
            "砂",
            "石",
            "沥青",
            "管",
            "电缆",
            "模板",
            "防水",
            "保温",
        )
        top_material_demand = sorted(
            [it for it in items if any(k in (it.name or "") for k in material_keywords)],
            key=lambda x: float(x.quantity or 0),
            reverse=True,
        )[:8]

        def _brief(it: BoQItem) -> dict[str, Any]:
            return {
                "boq_code": it.boq_code,
                "name": it.name,
                "quantity": it.quantity,
                "unit": it.unit,
                "unit_price": it.unit_price,
                "total_price": it.total_price,
            }

        hazard_keywords = ("危化", "易燃", "易爆", "有毒", "溶剂", "油漆", "涂料", "沥青", "气瓶")
        special_keywords = ("高性能", "特种", "抗渗", "防腐", "不锈钢", "高强", "预应力", "四新")
        ppe_keywords = ("安全帽", "安全带", "防护服", "防毒面具", "护目镜", "绝缘手套", "劳保")

        hazard_items = [it for it in items if any(k in (it.name or "") for k in hazard_keywords)][:12]
        special_items = [it for it in items if any(k in (it.name or "") for k in special_keywords)][:12]
        ppe_items = [it for it in items if any(k in (it.name or "") for k in ppe_keywords)][:12]
        chapter_structure = self._infer_chapter_structure(items)
        complexity_metrics = self._calc_complexity_metrics(items)

        output = {
            "total_quantity": total_qty,
            "item_count": count,
            "density": density,
            "top_quantity_items": [_brief(it) for it in top_quantity],
            "top_unit_price_items": [_brief(it) for it in top_unit_price],
            "top_total_price_items": [_brief(it) for it in top_total_price],
            "top_material_demand_items": [_brief(it) for it in top_material_demand],
            "special_material_items": [_brief(it) for it in special_items],
            "hazardous_material_items": [_brief(it) for it in hazard_items],
            "ppe_items": [_brief(it) for it in ppe_items],
            "chapter_structure": chapter_structure,
        }
        output.update(complexity_metrics)
        return output

    def _to_float(self, v) -> float | None:
        if v is None:
            return None
        s = str(v)
        s = re.sub(r"[^\d.]+", "", s)
        try:
            return float(s)
        except ValueError:
            return None
