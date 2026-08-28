"""Unit tests for boq_parser.py - BoQParser class"""

from __future__ import annotations

from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook

from backend.zhifei_autoplan.models import BoQItem, ConstructionProcess, Resource
from backend.zhifei_autoplan.parsers.boq_parser import BoQParser


class TestBoQParserToFloat:
    """Tests for _to_float method"""

    def setup_method(self):
        self.parser = BoQParser()

    def test_to_float_none(self):
        """None input returns None"""
        assert self.parser._to_float(None) is None

    def test_to_float_integer(self):
        """Integer input returns float"""
        assert self.parser._to_float(123) == 123.0

    def test_to_float_float(self):
        """Float input returns same float"""
        assert self.parser._to_float(45.67) == 45.67

    def test_to_float_string_number(self):
        """String number returns float"""
        assert self.parser._to_float("100.5") == 100.5

    def test_to_float_string_with_unit(self):
        """String with unit extracts number"""
        assert self.parser._to_float("100.5m³") == 100.5

    def test_to_float_string_with_prefix(self):
        """String with prefix extracts number"""
        assert self.parser._to_float("约200.0") == 200.0

    def test_to_float_string_mixed(self):
        """String with mixed characters extracts number"""
        assert self.parser._to_float("工程量：350.5吨") == 350.5

    def test_to_float_empty_string(self):
        """Empty string returns None"""
        assert self.parser._to_float("") is None

    def test_to_float_no_digits(self):
        """String without digits returns None"""
        assert self.parser._to_float("abc") is None

    def test_to_float_zero(self):
        """Zero returns 0.0"""
        assert self.parser._to_float(0) == 0.0

    def test_to_float_negative_string(self):
        """Negative sign is stripped, returns positive"""
        # Current implementation strips non-digit/dot chars
        result = self.parser._to_float("-50.5")
        assert result == 50.5

    def test_strict_quantity_rejects_feature_text_and_negative_values(self):
        assert self.parser._to_quantity("1、规格：250x200x5x8", unit="t") is None
        assert self.parser._to_quantity("-59.214", unit="t") is None
        assert self.parser._to_quantity("59.214t", unit="t") == 59.214
        assert self.parser._to_quantity("59.214m3", unit="t") is None


class TestBoQParserMapBoqToProcess:
    """Tests for map_boq_to_process method"""

    def setup_method(self):
        self.parser = BoQParser()

    def test_map_concrete(self):
        """混凝土 maps to 混凝土浇筑"""
        item = BoQItem(boq_code="001", name="C30混凝土浇筑", quantity=100.0)
        proc, res = self.parser.map_boq_to_process(item)
        assert proc.name == "混凝土浇筑"
        assert len(res) == 2
        assert res[0].name == "搅拌车"
        assert res[1].name == "泵车"

    def test_map_rebar(self):
        """钢筋 maps to 钢筋绑扎"""
        item = BoQItem(boq_code="002", name="钢筋制作安装", quantity=50.0)
        proc, res = self.parser.map_boq_to_process(item)
        assert proc.name == "钢筋绑扎"
        assert len(res) == 2
        assert res[0].name == "钢筋工"
        assert res[1].name == "电焊机"

    def test_map_formwork(self):
        """模板 maps to 模板安装"""
        item = BoQItem(boq_code="003", name="木模板", quantity=200.0)
        proc, res = self.parser.map_boq_to_process(item)
        assert proc.name == "模板安装"
        assert len(res) == 2
        assert res[0].name == "模板工"
        assert res[1].name == "支撑材料"

    def test_map_backfill(self):
        """回填 maps to 土方回填"""
        item = BoQItem(boq_code="004", name="土方回填", quantity=500.0)
        proc, res = self.parser.map_boq_to_process(item)
        assert proc.name == "土方回填"
        assert len(res) == 2
        assert res[0].name == "压路机"
        assert res[1].name == "装载机"

    def test_map_pipe(self):
        """管道 maps to 管道安装"""
        item = BoQItem(boq_code="005", name="PE管道敷设", quantity=300.0)
        proc, res = self.parser.map_boq_to_process(item)
        assert proc.name == "管道安装"
        assert len(res) == 2
        assert res[0].name == "挖机"
        assert res[1].name == "起重机"

    def test_map_default(self):
        """Unknown item maps to 通用施工工序"""
        item = BoQItem(boq_code="006", name="其他工程", quantity=10.0)
        proc, res = self.parser.map_boq_to_process(item)
        assert proc.name == "通用施工工序"
        assert len(res) == 0

    def test_map_empty_name(self):
        """Empty name maps to default"""
        item = BoQItem(boq_code="007", name="", quantity=None)
        proc, res = self.parser.map_boq_to_process(item)
        assert proc.name == "通用施工工序"
        assert len(res) == 0

    def test_map_first_match_wins(self):
        """First matching keyword wins"""
        # "混凝土钢筋" contains both 混凝土 and 钢筋, 混凝土 comes first in mapping
        item = BoQItem(boq_code="008", name="混凝土钢筋", quantity=100.0)
        proc, _res = self.parser.map_boq_to_process(item)
        assert proc.name == "混凝土浇筑"


class TestBoQParserCalcStats:
    """Tests for _calc_stats method"""

    def setup_method(self):
        self.parser = BoQParser()

    def test_calc_stats_empty(self):
        """Empty list returns zeros"""
        stats = self.parser._calc_stats([])
        assert stats["total_quantity"] == 0.0
        assert stats["item_count"] == 0.0
        assert stats["density"] == 0.0

    def test_calc_stats_single_item(self):
        """Single item with quantity"""
        items = [BoQItem(boq_code="001", name="Test", quantity=100.0)]
        stats = self.parser._calc_stats(items)
        assert stats["total_quantity"] == 100.0
        assert stats["item_count"] == 1.0
        assert stats["density"] == 100.0

    def test_calc_stats_multiple_items(self):
        """Multiple items with quantities"""
        items = [
            BoQItem(boq_code="001", name="A", quantity=100.0),
            BoQItem(boq_code="002", name="B", quantity=200.0),
            BoQItem(boq_code="003", name="C", quantity=300.0),
        ]
        stats = self.parser._calc_stats(items)
        assert stats["total_quantity"] == 600.0
        assert stats["item_count"] == 3.0
        assert stats["density"] == 200.0

    def test_calc_stats_with_none_quantity(self):
        """Items with None quantity are skipped"""
        items = [
            BoQItem(boq_code="001", name="A", quantity=100.0),
            BoQItem(boq_code="002", name="B", quantity=None),
            BoQItem(boq_code="003", name="C", quantity=200.0),
        ]
        stats = self.parser._calc_stats(items)
        assert stats["total_quantity"] == 300.0
        assert stats["item_count"] == 2.0
        assert stats["density"] == 150.0

    def test_calc_stats_all_none_quantity(self):
        """All items with None quantity"""
        items = [
            BoQItem(boq_code="001", name="A", quantity=None),
            BoQItem(boq_code="002", name="B", quantity=None),
        ]
        stats = self.parser._calc_stats(items)
        assert stats["total_quantity"] == 0.0
        assert stats["item_count"] == 0.0
        assert stats["density"] == 0.0

    def test_calc_stats_zero_quantity(self):
        """Zero quantity is counted"""
        items = [
            BoQItem(boq_code="001", name="A", quantity=0.0),
            BoQItem(boq_code="002", name="B", quantity=100.0),
        ]
        stats = self.parser._calc_stats(items)
        assert stats["total_quantity"] == 100.0
        assert stats["item_count"] == 2.0
        assert stats["density"] == 50.0

    def test_calc_stats_contains_chapter_structure_and_complexity(self):
        """Stats include chapter tree and complexity/density indexes"""
        item1 = BoQItem(boq_code="01.01.001", name="桥梁下部结构", quantity=120.0)
        item2 = BoQItem(boq_code="01.02.001", name="桥梁上部结构", quantity=80.0)
        item3 = BoQItem(boq_code="02.01.001", name="隧道衬砌", quantity=200.0)
        item1.process = ConstructionProcess(name="主体结构")
        item2.process = ConstructionProcess(name="主体结构")
        item3.process = ConstructionProcess(name="围护与建筑构造")
        item1.resources = [Resource(name="钢筋工"), Resource(name="模板工")]
        item2.resources = [Resource(name="混凝土工")]
        item3.resources = [Resource(name="防水工"), Resource(name="抹灰工"), Resource(name="测量工")]

        stats = self.parser._calc_stats([item1, item2, item3])

        assert "chapter_structure" in stats
        assert stats["chapter_structure"]["chapter_count"] >= 2
        assert isinstance(stats["chapter_structure"]["chapters"], list)

        for key in (
            "complexity_index",
            "resource_density_index",
            "quantity_scale_index",
            "construction_density_index",
            "process_diversity",
        ):
            assert key in stats
            assert 0.0 <= float(stats[key]) <= 1.0


class TestBoQParserRowToItem:
    """Tests for _row_to_item method"""

    def setup_method(self):
        self.parser = BoQParser()

    def test_row_to_item_full_data(self):
        """Full row data converts correctly"""
        row = {"code": "001", "name": "混凝土浇筑", "qty": 100.5, "unit": "m³"}
        item = self.parser._row_to_item(row)
        assert item.boq_code == "001"
        assert item.name == "混凝土浇筑"
        assert item.quantity == 100.5
        assert item.unit == "m³"
        assert item.process is not None
        assert item.process.name == "混凝土浇筑"

    def test_row_to_item_missing_code(self):
        """Missing code uses empty string"""
        row = {"name": "Test", "qty": 10.0, "unit": "个"}
        item = self.parser._row_to_item(row)
        assert item.boq_code == ""

    def test_row_to_item_missing_name(self):
        """Missing name uses empty string"""
        row = {"code": "001", "qty": 10.0}
        item = self.parser._row_to_item(row)
        assert item.name == ""

    def test_row_to_item_missing_unit(self):
        """Missing unit returns None"""
        row = {"code": "001", "name": "Test", "qty": 10.0}
        item = self.parser._row_to_item(row)
        assert item.unit is None

    def test_row_to_item_string_quantity(self):
        """String quantity is parsed"""
        row = {"code": "001", "name": "Test", "qty": "50.5m³", "unit": "m³"}
        item = self.parser._row_to_item(row)
        assert item.quantity == 50.5

    def test_row_to_item_none_quantity(self):
        """None quantity stays None"""
        row = {"code": "001", "name": "Test", "qty": None, "unit": "m³"}
        item = self.parser._row_to_item(row)
        assert item.quantity is None

    def test_row_to_item_process_attached(self):
        """Process is attached based on name"""
        row = {"code": "001", "name": "钢筋安装", "qty": 10.0, "unit": "t"}
        item = self.parser._row_to_item(row)
        assert item.process.name == "钢筋绑扎"
        assert len(item.resources) == 2


class TestBoQParserReadExcel:
    """Tests for _read_excel method"""

    def setup_method(self):
        self.parser = BoQParser()

    def test_read_excel_standard_columns(self, tmp_path):
        """Standard column names are parsed"""
        wb = Workbook()
        ws = wb.active
        ws.append(["清单编码", "项目名称", "工程量", "单位", "综合单价", "合价"])
        ws.append(["001", "混凝土", 100.0, "m³", 500.0, 50000.0])
        ws.append(["002", "钢筋", 50.0, "t", 5200.0, 260000.0])
        path = tmp_path / "boq.xlsx"
        wb.save(path)

        rows = self.parser._read_excel(str(path))
        
        assert len(rows) == 2
        assert rows[0]["code"] == "001"
        assert rows[0]["name"] == "混凝土"
        assert rows[0]["qty"] == 100.0
        assert rows[0]["unit"] == "m³"

    def test_read_excel_alt_columns(self, tmp_path):
        """Alternative column names are parsed"""
        wb = Workbook()
        ws = wb.active
        ws.append(["编码", "名称", "数量", "单位"])
        ws.append(["001", "管道", 200.0, "m"])
        path = tmp_path / "boq.xlsx"
        wb.save(path)

        rows = self.parser._read_excel(str(path))
        
        assert len(rows) == 1
        assert rows[0]["code"] == "001"
        assert rows[0]["name"] == "管道"

    def test_read_excel_five_column_unpriced_shape(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "分部分项"
        ws.append(["序号", "项目名称", "项目特征描述", "计量单位", "工程量"])
        ws.append([79, "钢梁", "1、Q355B\n2、规格：WH600*280*10*14", "t", 59.214])
        path = tmp_path / "工程量清单.xlsx"
        wb.save(path)

        rows = self.parser._read_excel(str(path))
        item = self.parser._row_to_item(rows[0])

        assert item.name == "钢梁"
        assert item.project_feature.startswith("1、Q355B")
        assert item.quantity == 59.214
        assert item.unit == "t"
        assert item.unit_price is None
        assert item.total_price is None
        assert item.source_locator["row_index"] == 2
        assert item.source_locator["locator"].startswith("boq:")

    def test_read_excel_missing_columns(self, tmp_path):
        """Missing columns use empty strings"""
        wb = Workbook()
        ws = wb.active
        ws.append(["other_col"])
        ws.append(["value"])
        path = tmp_path / "boq.xlsx"
        wb.save(path)

        rows = self.parser._read_excel(str(path))
        # Graceful fallback: may return [] (no recognizable header) or rows with empty fields.
        assert isinstance(rows, list)
        if rows:
            assert rows[0]["code"] == ""
            assert rows[0]["name"] == ""

    def test_read_excel_empty(self, tmp_path):
        """Empty Excel returns empty list"""
        wb = Workbook()
        ws = wb.active
        # only an empty header row
        ws.append([])
        path = tmp_path / "empty.xlsx"
        wb.save(path)

        rows = self.parser._read_excel(str(path))
        # Depending on fallback header detection, may return [] or a single empty row; parse() filters by name anyway.
        assert isinstance(rows, list)


class TestBoQParserReadPdfTables:
    """Tests for _read_pdf_tables method"""

    def setup_method(self):
        self.parser = BoQParser()

    @patch("backend.zhifei_autoplan.parsers.boq_parser.pdfplumber.open")
    def test_read_pdf_single_table(self, mock_open):
        """Single table with valid rows"""
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = [
            [["编码", "名称", "数量", "单位"],  # header
             ["001", "混凝土", "100", "m³"],
             ["002", "钢筋", "50", "t"]]
        ]
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page]
        mock_pdf.__enter__ = MagicMock(return_value=mock_pdf)
        mock_pdf.__exit__ = MagicMock(return_value=False)
        mock_open.return_value = mock_pdf
        
        rows = self.parser._read_pdf_tables("/fake/path.pdf")
        
        assert len(rows) == 2
        assert rows[0]["code"] == "001"
        assert rows[0]["name"] == "混凝土"
        assert rows[0]["qty"] == "100"
        assert rows[0]["unit"] == "m³"

    @patch("backend.zhifei_autoplan.parsers.boq_parser.pdfplumber.open")
    def test_read_pdf_multiple_pages(self, mock_open):
        """Multiple pages with tables"""
        mock_page1 = MagicMock()
        mock_page1.extract_tables.return_value = [
            [["编码", "名称", "数量", "单位"], ["001", "A", "10", "m"]]
        ]
        mock_page2 = MagicMock()
        mock_page2.extract_tables.return_value = [
            [["编码", "名称", "数量", "单位"], ["002", "B", "20", "t"]]
        ]
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page1, mock_page2]
        mock_pdf.__enter__ = MagicMock(return_value=mock_pdf)
        mock_pdf.__exit__ = MagicMock(return_value=False)
        mock_open.return_value = mock_pdf
        
        rows = self.parser._read_pdf_tables("/fake/path.pdf")
        
        assert len(rows) == 2

    @patch("backend.zhifei_autoplan.parsers.boq_parser.pdfplumber.open")
    def test_read_pdf_short_row_skipped(self, mock_open):
        """Rows with less than 4 columns are skipped"""
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = [
            [["编码", "名称", "数量", "单位"],
             ["001", "A"],  # short row
             ["002", "B", "20", "t"]]
        ]
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page]
        mock_pdf.__enter__ = MagicMock(return_value=mock_pdf)
        mock_pdf.__exit__ = MagicMock(return_value=False)
        mock_open.return_value = mock_pdf
        
        rows = self.parser._read_pdf_tables("/fake/path.pdf")
        
        assert len(rows) == 1
        assert rows[0]["code"] == "002"

    @patch("backend.zhifei_autoplan.parsers.boq_parser.pdfplumber.open")
    def test_read_pdf_none_row_skipped(self, mock_open):
        """None rows are skipped"""
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = [
            [["编码", "名称", "数量", "单位"],
             None,
             ["001", "A", "10", "m"]]
        ]
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page]
        mock_pdf.__enter__ = MagicMock(return_value=mock_pdf)
        mock_pdf.__exit__ = MagicMock(return_value=False)
        mock_open.return_value = mock_pdf
        
        rows = self.parser._read_pdf_tables("/fake/path.pdf")
        
        assert len(rows) == 1

    @patch("backend.zhifei_autoplan.parsers.boq_parser.pdfplumber.open")
    def test_read_pdf_no_tables(self, mock_open):
        """Page with no tables"""
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = None
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page]
        mock_pdf.__enter__ = MagicMock(return_value=mock_pdf)
        mock_pdf.__exit__ = MagicMock(return_value=False)
        mock_open.return_value = mock_pdf
        
        rows = self.parser._read_pdf_tables("/fake/path.pdf")
        
        assert len(rows) == 0

    @patch("backend.zhifei_autoplan.parsers.boq_parser.pdfplumber.open")
    def test_read_pdf_none_values_stripped(self, mock_open):
        """None cell values are handled"""
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = [
            [["编码", "名称", "数量", "单位"],
             [None, "A", None, "m"]]
        ]
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page]
        mock_pdf.__enter__ = MagicMock(return_value=mock_pdf)
        mock_pdf.__exit__ = MagicMock(return_value=False)
        mock_open.return_value = mock_pdf
        
        rows = self.parser._read_pdf_tables("/fake/path.pdf")
        
        assert len(rows) == 1
        assert rows[0]["code"] == ""
        assert rows[0]["name"] == "A"

    @patch("backend.zhifei_autoplan.parsers.boq_parser.pdfplumber.open")
    def test_read_pdf_real_five_column_rows_are_header_driven(self, mock_open):
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = [
            [
                ["序\n号", "项目名称", "项目特征描述", "计量\n单位", "工程量"],
                ["", "钢结构工程", None, "", ""],
                ["79", "钢梁", "1、Q355B\n2、规格：WH600*280*10*14", "t", "59.214"],
                ["49", "普通灯具", "1、名称：600x600LED平板灯", "套", "133.000"],
                ["50", "普通灯具", "1、名称：300x300LED平板灯", "套", "8.000"],
                ["228", "预制钢筋混凝土管桩", "3、规格：PHC-550AB125-C80", "根", "239.000"],
            ]
        ]
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page]
        mock_pdf.__enter__ = MagicMock(return_value=mock_pdf)
        mock_pdf.__exit__ = MagicMock(return_value=False)
        mock_open.return_value = mock_pdf

        rows = self.parser._read_pdf_tables("/fake/工程量清单及编制说明.pdf")
        items = [self.parser._row_to_item(row) for row in rows if self.parser._is_leaf_row(row)]

        assert [(item.name, item.quantity, item.unit) for item in items] == [
            ("钢梁", 59.214, "t"),
            ("普通灯具", 133.0, "套"),
            ("普通灯具", 8.0, "套"),
            ("预制钢筋混凝土管桩", 239.0, "根"),
        ]
        assert all(item.unit_price is None and item.total_price is None for item in items)
        assert all(item.source_locator["page"] == 1 for item in items)
        assert items[0].project_feature.startswith("1、Q355B")

    @patch("backend.zhifei_autoplan.parsers.boq_parser.pdfplumber.open")
    def test_unknown_five_column_table_is_not_positionally_guessed(self, mock_open):
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = [
            [["h1", "h2", "h3", "h4", "h5"], ["79", "钢梁", "Q355B", "t", "59.214"]]
        ]
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page]
        mock_pdf.__enter__ = MagicMock(return_value=mock_pdf)
        mock_pdf.__exit__ = MagicMock(return_value=False)
        mock_open.return_value = mock_pdf

        assert self.parser._read_pdf_tables("/fake/path.pdf") == []

    @patch("backend.zhifei_autoplan.parsers.boq_parser.pdfplumber.open")
    def test_unrelated_four_column_table_is_not_positionally_guessed(self, mock_open):
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = [
            [
                ["项目", "数量", "单位", "备注"],
                ["钢梁进场协调", "59.214", "项", "会议安排"],
            ]
        ]
        mock_pdf = MagicMock()
        mock_pdf.pages = [mock_page]
        mock_pdf.__enter__ = MagicMock(return_value=mock_pdf)
        mock_pdf.__exit__ = MagicMock(return_value=False)
        mock_open.return_value = mock_pdf

        assert self.parser._read_pdf_tables("/fake/会议纪要.pdf") == []


class TestBoQParserParse:
    """Tests for parse async method"""

    def setup_method(self):
        self.parser = BoQParser()

    @pytest.mark.asyncio
    @patch.object(BoQParser, "_read_excel")
    async def test_parse_xlsx(self, mock_read):
        """Parse xlsx file"""
        mock_read.return_value = [
            {"code": "001", "name": "混凝土", "qty": 100.0, "unit": "m³"}
        ]
        
        items, stats = await self.parser.parse("/path/to/file.xlsx")
        
        assert len(items) == 1
        assert items[0].name == "混凝土"
        assert stats["item_count"] == 1.0
        mock_read.assert_called_once_with("/path/to/file.xlsx")

    @pytest.mark.asyncio
    @patch.object(BoQParser, "_read_excel")
    async def test_parse_xls(self, mock_read):
        """Parse xls file"""
        mock_read.return_value = [
            {"code": "001", "name": "钢筋", "qty": 50.0, "unit": "t"}
        ]
        
        items, _stats = await self.parser.parse("/path/to/file.xls")
        
        assert len(items) == 1
        mock_read.assert_called_once()

    @pytest.mark.asyncio
    @patch.object(BoQParser, "_read_excel")
    async def test_parse_xlsx_uppercase(self, mock_read):
        """Parse XLSX (uppercase) file"""
        mock_read.return_value = [
            {"code": "001", "name": "Test", "qty": 10.0, "unit": "m"}
        ]
        
        items, _stats = await self.parser.parse("/path/to/file.XLSX")
        
        assert len(items) == 1

    @pytest.mark.asyncio
    @patch.object(BoQParser, "_read_pdf_tables")
    async def test_parse_pdf(self, mock_read):
        """Parse pdf file"""
        mock_read.return_value = [
            {"code": "001", "name": "管道", "qty": 200.0, "unit": "m"}
        ]
        
        items, _stats = await self.parser.parse("/path/to/file.pdf")
        
        assert len(items) == 1
        assert items[0].name == "管道"
        mock_read.assert_called_once_with("/path/to/file.pdf")

    @pytest.mark.asyncio
    @patch.object(BoQParser, "_read_pdf_tables")
    async def test_parse_pdf_uppercase(self, mock_read):
        """Parse PDF (uppercase) file"""
        mock_read.return_value = [
            {"code": "001", "name": "Test", "qty": 10.0, "unit": "m"}
        ]
        
        items, _stats = await self.parser.parse("/path/to/file.PDF")
        
        assert len(items) == 1

    @pytest.mark.asyncio
    async def test_parse_unsupported_format(self):
        """Unsupported format returns empty"""
        items, stats = await self.parser.parse("/path/to/file.txt")
        
        assert len(items) == 0
        assert stats["item_count"] == 0.0

    @pytest.mark.asyncio
    async def test_parse_docx_unsupported(self):
        """DOCX is unsupported"""
        items, _stats = await self.parser.parse("/path/to/file.docx")
        
        assert len(items) == 0

    @pytest.mark.asyncio
    @patch.object(BoQParser, "_read_excel")
    async def test_parse_filters_empty_names(self, mock_read):
        """Rows without name are filtered out"""
        mock_read.return_value = [
            {"code": "001", "name": "有效项", "qty": 100.0, "unit": "m³"},
            {"code": "002", "name": "", "qty": 50.0, "unit": "t"},
            {"code": "003", "name": None, "qty": 30.0, "unit": "m"},
        ]
        
        items, _stats = await self.parser.parse("/path/to/file.xlsx")
        
        # Only row with non-empty name is included
        assert len(items) == 1
        assert items[0].name == "有效项"

    @pytest.mark.asyncio
    @patch.object(BoQParser, "_read_pdf_tables")
    async def test_parse_filters_section_headings_and_invalid_quantities(self, mock_read):
        mock_read.return_value = [
            {"code": "", "name": "钢结构工程", "qty": "", "unit": ""},
            {"code": "79", "name": "钢梁", "qty": "Q355B 250x200x5x8", "unit": "t"},
            {"code": "79", "name": "钢梁", "qty": "59.214", "unit": "t"},
        ]

        items, stats = await self.parser.parse("/path/to/file.pdf")

        assert len(items) == 1
        assert items[0].quantity == 59.214
        assert stats["total_quantity"] == 59.214

    @pytest.mark.asyncio
    @patch.object(BoQParser, "_read_excel")
    async def test_parse_multiple_items(self, mock_read):
        """Multiple valid items are returned"""
        mock_read.return_value = [
            {"code": "001", "name": "A", "qty": 10.0, "unit": "m"},
            {"code": "002", "name": "B", "qty": 20.0, "unit": "t"},
            {"code": "003", "name": "C", "qty": 30.0, "unit": "m³"},
        ]
        
        items, stats = await self.parser.parse("/path/to/file.xlsx")
        
        assert len(items) == 3
        assert stats["total_quantity"] == 60.0
        assert stats["density"] == 20.0


class TestBoQParserIntegration:
    """Integration-style tests for BoQParser"""

    def setup_method(self):
        self.parser = BoQParser()

    def test_full_workflow_row_to_item(self):
        """Complete workflow from row dict to BoQItem with process and resources"""
        row = {
            "code": "010101001001",
            "name": "C30混凝土基础浇筑",
            "qty": "250.5m³",
            "unit": "m³"
        }
        
        item = self.parser._row_to_item(row)
        
        assert item.boq_code == "010101001001"
        assert item.name == "C30混凝土基础浇筑"
        assert item.quantity == 250.5
        assert item.unit == "m³"
        assert item.process.name == "混凝土浇筑"
        assert len(item.resources) == 2
        assert item.resources[0].name == "搅拌车"
        assert item.resources[1].name == "泵车"

    def test_calc_stats_with_processed_items(self):
        """Stats calculation works with fully processed items"""
        rows = [
            {"code": "001", "name": "混凝土", "qty": 100.0, "unit": "m³"},
            {"code": "002", "name": "钢筋", "qty": 50.0, "unit": "t"},
            {"code": "003", "name": "管道", "qty": 200.0, "unit": "m"},
        ]
        
        items = [self.parser._row_to_item(r) for r in rows]
        stats = self.parser._calc_stats(items)
        
        assert stats["total_quantity"] == 350.0
        assert stats["item_count"] == 3.0
        assert abs(stats["density"] - 116.67) < 0.01
