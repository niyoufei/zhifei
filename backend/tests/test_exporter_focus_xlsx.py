from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from backend.zhifei_autoplan.exporter import export_autoplan_focus_xlsx


def test_export_autoplan_focus_xlsx_includes_focus_card_columns(tmp_path: Path):
    sections = [
        {
            "title": "主要施工方法",
            "content": (
                "【清单重点项控制卡】\n"
                "- 清单项：钢筋；工程量=120t；单价=5200元/t；合价=624000元\n"
                "  量化指标：频次=1次/日；阈值=偏差≤5mm；间距=200mm；厚度=50mm；时长=4h/段；人数=8人/班；设备型号=20t挖机1台。\n"
                "  图纸定位：dwg.pdf#p1_12345678@90；校核点=构件位置/尺寸/标高/做法。【证据:dwg.pdf#p1_12345678@90】\n"
                "  标准引用：std.pdf#p2_abcdef12@34；条款对照入台账。【证据:std.pdf#p2_abcdef12@34】\n"
                "  风险→控制→验证：风险：误差累积导致返工；控制：放样复核=2次/日；验证：偏差≤5mm，记录=《复核记录》。【证据:boq.xlsx#p1_11111111@10】\n"
            ),
        }
    ]
    data = {
        "topic": "t1",
        "project_id": "p1",
        "sections": sections,
        "quality_checks": {"issue_list": [], "auto_revision_suggestions": []},
        "cross_index": {
            "project_id": "p1",
            "focus_count": 1,
            "mentioned_count": 1,
            "closed_ok_count": 1,
            "missing_drawing_locator_count": 0,
            "missing_standard_locator_count": 0,
            "focus_items": [
                {
                    "name": "钢筋",
                    "categories": ["工程量大"],
                    "process_name": "钢筋绑扎",
                    "boq_code": "010101",
                    "quantity": 120.0,
                    "unit": "t",
                    "unit_price": 5200.0,
                    "total_price": 624000.0,
                    "chapter": "主要施工方法",
                    "drawing_locator": "dwg.pdf#p1_12345678@90",
                    "standard_locator": "std.pdf#p2_abcdef12@34",
                    "closure": {"ok": True, "missing_parts": []},
                    "flags": [],
                    "evidence_locators_near": ["dwg.pdf#p1_12345678@90"],
                }
            ],
        },
    }
    out = tmp_path / "focus.xlsx"
    path = export_autoplan_focus_xlsx(data, str(out))
    assert path and Path(path).exists()

    wb = load_workbook(path)
    assert "focus_index" in wb.sheetnames
    assert "focus_cards" in wb.sheetnames

    ws = wb["focus_index"]
    headers = [c.value for c in ws[1]]
    assert "卡-频次" in headers
    assert "卡-风险" in headers
    assert "卡-证据来源" in headers

    col = {h: i + 1 for i, h in enumerate(headers) if h}
    # Row 2 is the first data row.
    assert ws.cell(row=2, column=col["卡-频次"]).value == "1次/日"
    assert "误差累积" in str(ws.cell(row=2, column=col["卡-风险"]).value or "")
    assert "dwg.pdf#p1_12345678@90" in str(ws.cell(row=2, column=col["卡-证据来源"]).value or "")


def test_export_autoplan_focus_xlsx_includes_variant_similarity_sheet(tmp_path: Path):
    data = {
        "topic": "t1",
        "project_id": "p1",
        "sections": [],
        "quality_checks": {"issue_list": [], "auto_revision_suggestions": []},
        "cross_index": {"project_id": "p1", "focus_items": []},
        "variant_similarity": {
            "ok": True,
            "variant_count": 3,
            "chapter_threshold": 0.9,
            "overall_threshold": 0.85,
            "min_chars": 800,
            "avg_max_similarity": 0.72,
            "flagged_count": 0,
            "flagged": [],
            "by_chapter": [
                {
                    "title": "主要施工方法",
                    "lens": [1200, 1180, 1250],
                    "max_pair": "v1_v2",
                    "max_combined": 0.72,
                    "v1_v2": {"combined": 0.72, "jaccard3": 0.7, "cosine2": 0.75},
                    "v1_v3": {"combined": 0.66, "jaccard3": 0.62, "cosine2": 0.7},
                    "v2_v3": {"combined": 0.64, "jaccard3": 0.6, "cosine2": 0.68},
                }
            ],
        },
    }
    out = tmp_path / "focus.xlsx"
    path = export_autoplan_focus_xlsx(data, str(out))
    assert path and Path(path).exists()
    wb = load_workbook(path)
    assert "variant_similarity" in wb.sheetnames
