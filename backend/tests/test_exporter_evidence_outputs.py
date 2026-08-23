from __future__ import annotations

from pathlib import Path

from docx import Document
from openpyxl import load_workbook
import pytest

from backend.zhifei_autoplan.exporter import (
    export_expert_review_brief_docx,
    export_scoring_evidence_overview_xlsx,
)


pytestmark = pytest.mark.usefixtures("allow_legacy_export_contract")


def test_export_scoring_evidence_overview_xlsx(tmp_path: Path):
    data = {
        "topic": "测试项目施工组织设计",
        "project_id": "P-001",
        "score_mapping": {
            "item_cards": [
                {
                    "item_id": "ITEM-001",
                    "dimension": "主要施工方法",
                    "keywords": ["施工方法", "工艺流程"],
                    "matched_keywords": ["施工方法"],
                    "missing_keywords": ["工艺流程"],
                    "coverage_ratio": 0.5,
                    "estimated_score": 42.0,
                    "deduction_risk": 0.6,
                    "matched_sections": ["主要施工方法"],
                }
            ],
            "high_risk_items": [{"item_id": "ITEM-001"}],
        },
        "evidence_tracking": {
            "rows": [
                {
                    "paragraph_id": "S01-P001",
                    "section_title": "主要施工方法",
                    "paragraph_index": 1,
                    "page_estimate": 8,
                    "tender_score_points": [
                        {"rule_id": "ITEM-001", "dimension": "主要施工方法", "matched_keywords": ["施工方法"]}
                    ],
                    "evidence_sources": ["图纸A.dxf#p1_abcdef12@20", "GB50300-2013 5.2.1"],
                    "evidence_typed": {
                        "graph_nodes": ["ZF-KG-08#$.章节.施工方法"],
                        "drawing_refs": ["图纸A.dxf#p1_abcdef12@20"],
                        "standard_refs": ["GB50300-2013 5.2.1"],
                        "other_refs": [],
                    },
                    "system_response": "本章描述关键施工方法与控制流程。",
                }
            ],
            "summary": {
                "paragraph_count": 1,
                "score_point_bound_rows": 1,
                "evidence_bound_rows": 1,
                "traceable_locator_rows": 1,
            },
        },
    }

    out = tmp_path / "score_overview.xlsx"
    path = export_scoring_evidence_overview_xlsx(data, str(out))
    assert path and Path(path).exists()

    wb = load_workbook(path)
    assert "summary" in wb.sheetnames
    assert "score_items" in wb.sheetnames
    assert "score_evidence_matrix" in wb.sheetnames
    assert "paragraph_evidence" in wb.sheetnames

    ws = wb["score_evidence_matrix"]
    headers = [c.value for c in ws[1]]
    col = {str(h): i + 1 for i, h in enumerate(headers) if h}
    assert ws.cell(row=2, column=col["score_rule_id"]).value == "ITEM-001"
    assert ws.cell(row=2, column=col["page_estimate"]).value == 8


def test_export_expert_review_brief_docx(tmp_path: Path):
    data = {
        "topic": "测试项目施工组织设计",
        "style": {"body_font": "宋体", "title_font": "宋体"},
        "sections": [
            {"title": "施工进度计划", "content": "关键线路控制，加分策略触发。"},
            {"title": "安全保证措施", "content": "重大风险闭环：风险-控制-验证。"},
        ],
        "quality_checks": {
            "issue_list": [
                {
                    "severity": "high",
                    "title": "安全保证措施",
                    "type": "risk_triplet_missing",
                    "problem": "风险闭环不完整",
                    "suggestion": "补充验证与记录频次",
                }
            ]
        },
        "boq_wbs_cpm": {
            "summary": {
                "project_duration_days": 320,
                "resource_peak": 180,
                "critical_interval_days": 7,
                "critical_path_names": ["测量放线", "基础施工", "主体结构"],
            },
            "wbs": [{"process_name": "测量放线"}],
        },
    }
    out = tmp_path / "expert_review.docx"
    path = export_expert_review_brief_docx(data, str(out))
    assert path and Path(path).exists()

    doc = Document(path)
    text = "\n".join([p.text for p in doc.paragraphs if p.text])
    assert "专家复核提要版" in text
    assert "资源峰值" in text
    assert "qt_score_booster" in text or "加分策略" in text
