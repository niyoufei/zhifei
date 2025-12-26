# backend/export_excel.py
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from backend.m9_report_schema import ReportBundle

def export_report_to_excel(bundle: ReportBundle, filepath: str):
    """
    将 ReportBundle 导出为 Excel 报表：
    - Sheet1: 引用索引表
    - Sheet2: 评分点覆盖清单
    - Sheet3: 10%复核提要
    """
    wb = Workbook()

    # --- Sheet1: 引用索引表 ---
    ws1 = wb.active
    ws1.title = "引用索引表"
    ws1.append(["证据ID", "来源文件", "页码", "定位符", "摘录"])
    for ev in bundle.appendix.reference_index:
        ws1.append([ev.evidence_id, ev.source_name, ev.page, ev.locator, ev.snippet or ""])
    for col in range(1, 6):
        ws1.column_dimensions[get_column_letter(col)].width = 20

    # --- Sheet2: 评分点覆盖清单 ---
    ws2 = wb.create_sheet("评分点覆盖清单")
    ws2.append(["规则ID", "评分点", "得分", "满分", "通过", "备注", "关联证据"])
    for s in bundle.appendix.scoring_coverage:
        ws2.append([s.rule_id, s.title, s.score, s.max_score, "✅" if s.passed else "❌",
                    s.remarks or "", ", ".join(s.evidence_ids)])
    for col in range(1, 8):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # --- Sheet3: 10%复核提要 ---
    ws3 = wb.create_sheet("复核提要")
    ws3.append(["章节ID", "标题", "摘要内容"])
    for sec in bundle.sections:
        if sec.content_summary:
            ws3.append([sec.section_id, sec.title, sec.content_summary[:400]])
    ws3.append([])
    ws3.append(["导出时间", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])

    wb.save(filepath)
    return filepath

